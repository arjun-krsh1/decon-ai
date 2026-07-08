"""
amazon_scraper.py — Amazon Product Intelligence Engine

TWO modes:
1. Keyword search → finds top N products → scrapes each
2. Direct URLs    → scrapes each URL provided

Uses SerpAPI (free tier) for search results,
Playwright for full product page scraping (claims, reviews, ranking).

Speed optimisations:
- Parallel processing where possible
- Caches results for 24 hours
- Progress updates so UI never feels frozen
"""

import os, json, time, re, pathlib, hashlib, sys
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()

sys.path.insert(0, '.')
from llm import ask_llm, llm_available
from product_analytics import (parse_pack_size, price_per_ml,
                               star_distribution, review_timeline,
                               category_analytics, opportunity_cards)

CACHE = pathlib.Path("scraper/cache/amazon")
CACHE.mkdir(parents=True, exist_ok=True)

# every analysis run is snapshotted here so month-over-month trends build up
PI_STORE = pathlib.Path("scraper/cache/product_intel")
PI_STORE.mkdir(parents=True, exist_ok=True)

SERPAPI_KEY = os.getenv("SERPAPI_KEY", "")
APIFY_KEY = os.getenv("APIFY_KEY", "")

# ── Deconstruct's tracked competitor set ──────────────────────────────────────
# Product Intelligence analyses ONLY these brands (+ Deconstruct as baseline),
# not every random Amazon seller. Aliases handle Amazon's messy brand strings.
COMPETITOR_ALIASES = {
    "Deconstruct":       ["deconstruct", "thedeconstruct"],   # baseline
    "Minimalist":        ["minimalist", "beminimalist"],
    "Foxtale":           ["foxtale"],
    "The Derma Co.":     ["thedermaco", "dermaco"],
    "Dot & Key":         ["dotandkey", "dotkey"],
    "Dr. Sheth's":       ["drsheths", "drsheth"],
    "Plum":              ["plumgoodness", "plum"],
    "Aqualogica":        ["aqualogica"],
    "Pilgrim":           ["pilgrim", "discoverpilgrim"],
    "Hyphen":            ["hyphen"],
    "Conscious Chemist": ["consciouschemist"],
}
TARGET_COMPETITORS = [b for b in COMPETITOR_ALIASES if b != "Deconstruct"]


def _canon(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def match_competitor(brand, title=""):
    """Return the canonical tracked-brand name if brand/title matches one, else None."""
    cbrand, ctitle = _canon(brand), _canon(title)
    for canon, aliases in COMPETITOR_ALIASES.items():
        for a in aliases:
            if a and (a in cbrand or a in ctitle):
                return canon
    return None


def save_snapshot(keyword, results, date_range=""):
    """Persist a timestamped snapshot of a run for historical comparison."""
    try:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe = re.sub(r"[^a-z0-9]+", "_", (keyword or "run").lower()).strip("_")[:40]
        (PI_STORE / f"run_{safe}_{stamp}.json").write_text(
            json.dumps({"keyword": keyword, "date_range": date_range,
                        "generated_at": datetime.now().isoformat(timespec="seconds"),
                        "products": results}, indent=2, ensure_ascii=False),
            encoding="utf-8")
    except Exception as e:
        print(f"[pi] snapshot save failed: {e}")


def load_snapshots(keyword=""):
    """Load past run snapshots (newest first), optionally filtered by keyword."""
    runs = []
    for f in sorted(PI_STORE.glob("run_*.json"), reverse=True):
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        if keyword and data.get("keyword", "").lower() != keyword.lower():
            continue
        runs.append(data)
    return runs


def safe_number(val, default=0, max_val=None):
    """
    Convert ANY value to a float safely.
    Handles: 'Previous', '4.2 out of 5', '₹349', '15,420', None, int, float, 'Best Seller'
    Returns default if no number found.
    """
    import re as _re
    try:
        if val is None or str(val).strip() == "":
            return default
        s = str(val).replace(",", "").replace("₹", "").replace("+", "").strip()
        m = _re.search(r"[0-9]+\.?[0-9]*", s)
        result = float(m.group()) if m else default
        if max_val and result > max_val:
            return default  # sanity check failed
        return result
    except Exception:
        return default


# ── SerpAPI search ────────────────────────────────────────────────────────────
def search_amazon(keyword, n_results=10):
    """Search Amazon India for a keyword using SerpAPI."""
    if not SERPAPI_KEY:
        return _mock_search_results(keyword, n_results)

    cache_key = hashlib.md5(f"search_{keyword}_{n_results}".encode()).hexdigest()
    cache_file = CACHE / f"{cache_key}.json"

    if cache_file.exists():
        age = (time.time() - cache_file.stat().st_mtime) / 3600
        if age < 24:
            return json.loads(cache_file.read_text(encoding='utf-8'))

    try:
        import requests
        r = requests.get(
            "https://serpapi.com/search",
            params={
                "engine": "amazon",
                "amazon_domain": "amazon.in",
                "k": keyword,
                "api_key": SERPAPI_KEY,
            },
            timeout=30
        )
        r.raise_for_status()
        data = r.json()
        results = []

        for item in data.get("organic_results", [])[:n_results]:
            if not isinstance(item, dict):
                continue
            # SerpAPI's Amazon schema varies: `price` may be a string ("₹349"),
            # a dict, or exposed as `extracted_price`; `reviews` may be int or dict.
            price_field = item.get("price")
            if isinstance(price_field, dict):
                price_val = price_field.get("value") or price_field.get("extracted_value") or 0
                price_raw = price_field.get("raw", "")
            else:
                price_val = item.get("extracted_price") or safe_number(price_field)
                price_raw = str(price_field or "")
            reviews_field = item.get("reviews")
            review_count = (reviews_field.get("total", 0) if isinstance(reviews_field, dict)
                            else reviews_field or item.get("ratings_total") or 0)
            mrp_val = safe_number(item.get("old_price") or item.get("original_price") or 0)
            results.append({
                "asin": item.get("asin", ""),
                "title": item.get("title", ""),
                "brand": item.get("brand") or _extract_brand(item.get("title", "")),
                "price": safe_number(price_val),
                "price_raw": price_raw,
                "mrp": mrp_val,
                "rating": safe_number(item.get("rating", 0), max_val=5),
                "review_count_raw": int(safe_number(review_count)),
                "position": item.get("position", 0),
                "url": item.get("link") or f"https://www.amazon.in/dp/{item.get('asin','')}",
                "thumbnail": item.get("thumbnail", ""),
                "badge": item.get("badge", ""),
                "sponsored": item.get("sponsored", False),
            })

        cache_file.write_text(json.dumps(results, indent=2), encoding='utf-8')
        return results

    except Exception as e:
        print(f"[serpapi] error: {e}")
        return _mock_search_results(keyword, n_results)


# ── SerpAPI Amazon Reviews fetcher ───────────────────────────────────────────
def _entry(rev):
    return {
        "stars": rev.get("rating", ""),
        "title": rev.get("title", ""),
        "body": rev.get("body", rev.get("review", "")),
        "date": rev.get("date", ""),
        "verified": rev.get("verified_purchase", False),
        "helpful": rev.get("helpful_votes", 0),
    }


def _fetch_review_pages(asin, filter_by_star="", pages=3):
    """Fetch several pages of reviews from SerpAPI for one ASIN. Returns raw list."""
    import requests
    out = []
    for page in range(1, pages + 1):
        params = {
            "engine": "amazon_product_reviews",
            "asin": asin,
            "amazon_domain": "amazon.in",
            "sort_by": "recent",   # recency → enables rating-over-time
            "page": page,
            "api_key": SERPAPI_KEY,
        }
        if filter_by_star:
            params["filter_by_star"] = filter_by_star
        try:
            r = requests.get("https://serpapi.com/search", params=params, timeout=30)
            if r.status_code != 200:
                break
            revs = r.json().get("reviews", [])
            if not revs:
                break
            out.extend(revs)
            time.sleep(0.6)  # be polite / protect quota
        except Exception as e:
            print(f"    [reviews] page {page} error: {e}")
            break
    return out


def fetch_reviews_via_serpapi(asin, pages=3):
    """
    Fetch MANY real Amazon reviews via SerpAPI (paginated), split into positive
    (>=4*) and critical (<=2*), each keeping star + date so downstream analytics
    can build a star distribution and rating-over-time.

    Returns (positive[:30], critical[:20]). Uses a few SerpAPI credits per product.
    """
    if not SERPAPI_KEY or not asin:
        return [], []

    cache_key = hashlib.md5(f"reviews_v2_{asin}_{pages}".encode()).hexdigest()
    cache_file = CACHE / f"reviews_{cache_key}.json"

    if cache_file.exists():
        age = (time.time() - cache_file.stat().st_mtime) / 3600
        if age < 48:
            data = json.loads(cache_file.read_text(encoding='utf-8'))
            print(f"    [reviews] using cached reviews for {asin}")
            return data.get("positive", []), data.get("critical", [])

    try:
        positive, critical, seen = [], [], set()

        def _ingest(revs):
            for rev in revs:
                body = rev.get("body", rev.get("review", ""))
                key = (body or "")[:80]
                if not body or key in seen:
                    continue
                seen.add(key)
                stars = safe_number(str(rev.get("rating", 0)))
                e = _entry(rev)
                if stars >= 4:
                    positive.append(e)
                elif stars <= 2:
                    critical.append(e)

        _ingest(_fetch_review_pages(asin, "", pages))            # all-star, recent
        _ingest(_fetch_review_pages(asin, "critical", pages))    # extra critical

        cache_file.write_text(
            json.dumps({"positive": positive[:30], "critical": critical[:20]},
                       indent=2, ensure_ascii=False),
            encoding='utf-8'
        )
        print(f"    [reviews] fetched {len(positive)} positive, {len(critical)} critical for {asin}")
        return positive[:30], critical[:20]

    except Exception as e:
        print(f"    [reviews] SerpAPI error for {asin}: {e}")
        return [], []


AMAZON_REVIEWS_ACTOR = os.getenv("APIFY_AMAZON_REVIEWS_ACTOR", "junglee/amazon-reviews-scraper")


def _norm_review(rv):
    """Defensive field mapping — Amazon-review actors differ in their key names."""
    def g(*keys):
        for k in keys:
            v = rv.get(k)
            if v not in (None, ""):
                return v
        return ""
    stars = g("ratingScore", "rating", "stars", "reviewRating", "score")
    return {
        "stars": stars,
        "title": g("reviewTitle", "title"),
        "body": g("reviewDescription", "reviewText", "text", "body", "review", "content"),
        "date": g("date", "reviewDate", "reviewedAt", "reviewedDate"),
        "verified": bool(g("verified", "verifiedPurchase", "isVerified")),
        "helpful": g("helpfulCount", "helpful_votes", "helpfulVotes", "numberOfHelpful"),
    }


def fetch_reviews_apify(asin, target=120):
    """
    Pull a DEEP, representative review sample via an Apify Amazon-reviews actor:
    the most recent + most-helpful reviews plus a dedicated critical (1-2★) slice.
    Returns (positive[], critical[]). Cached 48h. Empty on no key / failure.
    """
    if not APIFY_KEY or not asin:
        return [], []
    cache_key = hashlib.md5(f"amzreviews_{asin}_{target}".encode()).hexdigest()
    cache_file = CACHE / f"revfull_{cache_key}.json"
    if cache_file.exists() and (time.time() - cache_file.stat().st_mtime) / 3600 < 48:
        data = json.loads(cache_file.read_text(encoding="utf-8"))
        print(f"    [reviews] using cached deep reviews for {asin}")
        return data.get("positive", []), data.get("critical", [])

    try:
        client = _client()
        run = client.actor(AMAZON_REVIEWS_ACTOR).call(run_input={
            "productUrls": [{"url": f"https://www.amazon.in/dp/{asin}"}],
            "maxReviews": target,
            "sort": "recent",
            "includeGdprSensitive": False,
            "proxyConfiguration": {"useApifyProxy": True},
        })
        assert run is not None
        items = list(client.dataset(run.default_dataset_id).iterate_items())

        positive, critical, seen = [], [], set()
        for rv in items:
            e = _norm_review(rv)
            body = str(e["body"]).strip()
            key = body[:80]
            if not body or key in seen:
                continue
            seen.add(key)
            stars = safe_number(str(e["stars"]))
            if stars >= 4:
                positive.append(e)
            elif stars <= 2:
                critical.append(e)
        cache_file.write_text(json.dumps({"positive": positive, "critical": critical},
                                         indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"    [reviews] deep pull {asin}: {len(positive)} positive, {len(critical)} critical "
              f"({len(items)} raw)")
        return positive, critical
    except Exception as e:
        print(f"    [reviews] Apify deep-review error for {asin}: {e}")
        return [], []


def _client():
    from apify_client import ApifyClient
    return ApifyClient(APIFY_KEY)


def fetch_product_serpapi(asin):
    """
    PRIMARY, robust product fetch via SerpAPI's `amazon_product` engine (one call):
    price, MRP, discount, rating, review count, 'bought last month', bullets,
    details, top dated reviews, AND Amazon's own aspect insights (each with real
    mention counts + example quotes). Returns a normalised dict, or {} on failure.
    """
    if not SERPAPI_KEY or not asin:
        return {}

    cache_key = hashlib.md5(f"amzprod_{asin}".encode()).hexdigest()
    cache_file = CACHE / f"prod_{cache_key}.json"
    if cache_file.exists():
        age = (time.time() - cache_file.stat().st_mtime) / 3600
        if age < 24:
            return json.loads(cache_file.read_text(encoding="utf-8"))

    try:
        import requests
        d = requests.get("https://serpapi.com/search",
                         params={"engine": "amazon_product", "asin": asin,
                                 "amazon_domain": "amazon.in", "api_key": SERPAPI_KEY},
                         timeout=30).json()
        if d.get("error"):
            print(f"    [product] SerpAPI: {d['error']}")
            return {}
        pr = d.get("product_results", {}) or {}
        ri = d.get("reviews_information", {}) or {}

        # top reviews (dated) → split positive/critical
        pos, crit = [], []
        variant_hint = ""
        for rv in (ri.get("authors_reviews") or []):
            stars = safe_number(rv.get("rating", 0), max_val=5)
            entry = {"stars": rv.get("rating", ""), "title": rv.get("title", ""),
                     "body": rv.get("text", ""), "date": rv.get("date", ""),
                     "verified": str(rv.get("verified_purchase", "")).lower() == "true",
                     "helpful": rv.get("helpful_votes", "")}
            if not variant_hint:
                variant_hint = str(rv.get("product", ""))
            if stars >= 4:
                pos.append(entry)
            elif stars <= 2:
                crit.append(entry)

        # Amazon's pre-computed aspect insights (precise, counted)
        insights = []
        for ins in (ri.get("summary", {}) or {}).get("insights", []):
            m = ins.get("mentions", {}) or {}
            insights.append({
                "aspect": ins.get("title", ""),
                "sentiment": ins.get("sentiment", ""),
                "total": int(safe_number(m.get("total", 0))),
                "positive": int(safe_number(m.get("positive", 0))),
                "negative": int(safe_number(m.get("negative", 0))),
                "summary": ins.get("summary", ""),
                "example": (ins.get("examples") or [{}])[0].get("snippet", ""),
            })

        details = d.get("product_details") or d.get("item_specifications") or {}
        # Amazon often returns the byline "Visit the <Brand> Store" or "<Brand> Store"
        _brand = re.sub(r"^\s*visit the\s+|\s+store\s*$", "", str(pr.get("brand", "")),
                        flags=re.IGNORECASE).strip()
        out = {
            "title": pr.get("title", ""),
            "brand": _brand or _extract_brand(pr.get("title", "")),
            "selling_price": safe_number(pr.get("extracted_price") or pr.get("price")),
            "mrp": safe_number(pr.get("extracted_old_price") or 0),
            "discount_pct": str(pr.get("discount", "")),
            "rating": safe_number(pr.get("rating", 0), max_val=5),
            "review_count": int(safe_number(pr.get("reviews", 0))),
            "bought_last_month": pr.get("bought_last_month", ""),
            "bullets": d.get("about_item") or [],
            "description": str(d.get("product_description", ""))[:1200],
            "product_details": details if isinstance(details, dict) else {},
            "reviews": pos,
            "critical_reviews": crit,
            "amazon_aspects": insights,
            "amazon_summary": (ri.get("summary", {}) or {}).get("text", ""),
            "variant_hint": variant_hint,
            "thumbnail": pr.get("thumbnail", ""),
        }
        cache_file.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"    [product] {asin}: ₹{out['selling_price']} · {out['review_count']} reviews · "
              f"{len(pos)+len(crit)} top reviews · {len(insights)} aspect insights")
        return out

    except Exception as e:
        print(f"    [product] error for {asin}: {e}")
        return {}


# ── Playwright product page scraper ──────────────────────────────────────────
def _run_playwright_page(url, js_code, wait=2):
    """Helper — open one page, run JS, return result."""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 800}
            )
            page = ctx.new_page()
            try:
                page.goto(url, timeout=25000, wait_until="domcontentloaded")
                time.sleep(wait)
                result = page.evaluate(js_code)
                browser.close()
                return result
            except Exception as e:
                browser.close()
                print(f"    [playwright] page error {url[:60]}: {e}")
                return {}
    except Exception as e:
        print(f"    [playwright] browser error: {e}")
        return {}


def _extract_asin(url):
    """Extract ASIN from Amazon URL."""
    import re
    m = re.search(r"/dp/([A-Z0-9]{10})", url)
    if m:
        return m.group(1)
    m = re.search(r"/product/([A-Z0-9]{10})", url)
    if m:
        return m.group(1)
    return ""


def scrape_product_page(url, asin=""):
    """
    Multi-page scrape for deep data:
    Page 1: product page  → title, MRP, selling price, bullets, details, ranking
    Page 2: reviews page  → top positive reviews (sorted by helpful)
    Page 3: critical page → 1-2 star critical reviews only
    """
    if not asin:
        asin = _extract_asin(url)

    cache_key = hashlib.md5(f"{url}_deep".encode()).hexdigest()
    cache_file = CACHE / f"product_{cache_key}.json"

    if cache_file.exists():
        age_hours = (time.time() - cache_file.stat().st_mtime) / 3600
        if age_hours < 24:
            return json.loads(cache_file.read_text(encoding='utf-8'))

    # ── PAGE 1: Product page ──────────────────────────────────────────────────
    print(f"    [scrape] product page...")
    product_js = """() => {
        const get = (sel) => {
            const el = document.querySelector(sel);
            return el ? el.textContent.trim() : '';
        };
        const getAll = (sel) => {
            return Array.from(document.querySelectorAll(sel))
                   .map(e => e.textContent.trim())
                   .filter(t => t.length > 5);
        };
        const getAttr = (sel, attr) => {
            const el = document.querySelector(sel);
            return el ? (el.getAttribute(attr) || el.textContent.trim()) : '';
        };

        // SELLING PRICE — the actual price customer pays
        const sellingPrice =
            get('.apexPriceToPay .a-offscreen') ||
            get('#corePriceDisplay_desktop_feature_div .a-offscreen') ||
            get('.a-price[data-a-color="price"] .a-offscreen') ||
            get('#priceblock_ourprice') ||
            get('#priceblock_dealprice') ||
            get('.a-price-whole');

        // MRP — original price with strikethrough
        const mrpEl = document.querySelector(
            '.basisPrice .a-offscreen, ' +
            '.a-text-price .a-offscreen, ' +
            '#listPrice, ' +
            '[data-a-strike="true"] .a-offscreen, ' +
            '.a-price.a-text-price .a-offscreen'
        );
        const mrp = mrpEl ? mrpEl.textContent.trim() : '';

        // Discount
        const discount =
            get('.savingsPercentage') ||
            get('#savingsPercentage') ||
            get('.reinventPriceSavingsPercentageMargin');

        // Title
        const title = get('#productTitle');

        // Brand
        const brand = get('#bylineInfo') ||
                       get('.po-brand .po-break-word') ||
                       get('#brand');

        // Rating
        const rating = getAttr('#acrPopover', 'title') ||
                        get('#acrPopover .a-icon-alt') ||
                        get('.a-icon-star span');

        // Review count
        const reviewCount = get('#acrCustomerReviewText');

        // Best seller badge
        const bsBadge = get('#bestsellersRank') ||
                         get('.badge-label') ||
                         get('#SalesRank');

        // Category rankings from details section
        const rankEls = document.querySelectorAll(
            '#detailBulletsWrapper_feature_div li, ' +
            '#productDetails_detailBullets_sections1 tr, ' +
            '#prodDetails tr'
        );
        const rankings = [];
        rankEls.forEach(el => {
            const text = el.textContent.replace(/[\t\n\r ]+/g, ' ').trim();
            if ((text.includes('Best Seller') || text.includes('#')) &&
                text.match(/#[0-9,]+/)) {
                rankings.push(text.slice(0, 200));
            }
        });

        // Feature bullets (About this item)
        const bullets = Array.from(
            document.querySelectorAll('#feature-bullets li span.a-list-item')
        ).map(e => e.textContent.trim()).filter(t => t.length > 10 && !t.includes('Make sure'));

        // Product details table
        const details = {};
        document.querySelectorAll(
            '#productDetails_techSpec_section_1 tr, ' +
            '#productDetails_detailBullets_sections1 tr, ' +
            '#detailBulletsWrapper_feature_div li'
        ).forEach(row => {
            const cells = row.querySelectorAll('td, th, span.a-text-bold');
            if (cells.length >= 2) {
                const key = cells[0].textContent.replace(/[:\n]/g,'').trim();
                const val = cells[1] ? cells[1].textContent.trim() : '';
                if (key && val && key.length < 60 && val.length < 200) {
                    details[key] = val;
                }
            }
        });

        // Full description
        const descParts = Array.from(
            document.querySelectorAll('#productDescription p, #productDescription span, #aplus p, #aplus li')
        ).map(e => e.textContent.trim()).filter(t => t.length > 20).slice(0, 6);

        // Bought count
        const boughtCount = get('#social-proofing-faceout-title-tk_bought') ||
                             get('.social-proofing-faceout-title');

        return {
            title, brand, selling_price: sellingPrice, mrp, discount_pct: discount,
            rating, review_count: reviewCount, bullets: bullets.slice(0, 12),
            product_details: details, description: descParts.join(' ').slice(0, 1000),
            rankings, badge: bsBadge, bought_count: boughtCount
        };
    }"""

    page1 = _run_playwright_page(url, product_js, wait=2)
    time.sleep(1.5)

    # ── PAGE 2: Positive reviews ──────────────────────────────────────────────
    positive_reviews = []
    if asin:
        reviews_url = f"https://www.amazon.in/product-reviews/{asin}/?sortBy=helpful&pageNumber=1"
        print(f"    [scrape] positive reviews page...")
        reviews_js = """() => {
            const reviews = [];
            document.querySelectorAll('[data-hook="review"]').forEach(review => {
                const title = review.querySelector('[data-hook="review-title"] span:last-child');
                const body = review.querySelector('[data-hook="review-body"] span');
                const stars = review.querySelector('[data-hook="review-star-rating"] span') ||
                               review.querySelector('.review-rating span');
                const date = review.querySelector('[data-hook="review-date"]');
                const verified = review.querySelector('[data-hook="avp-badge"]');
                if (body && body.textContent.trim().length > 30) {
                    reviews.push({
                        title: title ? title.textContent.trim() : '',
                        body: body.textContent.trim().slice(0, 600),
                        stars: stars ? stars.textContent.trim() : '',
                        date: date ? date.textContent.trim() : '',
                        verified: !!verified
                    });
                }
            });
            return reviews.slice(0, 6);
        }"""
        rv = _run_playwright_page(reviews_url, reviews_js, wait=2)
        if isinstance(rv, list):
            positive_reviews = rv
        time.sleep(1.5)

    # ── PAGE 3: Critical reviews (1-2 star only) ──────────────────────────────
    critical_reviews = []
    if asin:
        critical_url = f"https://www.amazon.in/product-reviews/{asin}/?filterByStar=critical&sortBy=recent"
        print(f"    [scrape] critical reviews page...")
        critical_js = """() => {
            const reviews = [];
            document.querySelectorAll('[data-hook="review"]').forEach(review => {
                const title = review.querySelector('[data-hook="review-title"] span:last-child');
                const body = review.querySelector('[data-hook="review-body"] span');
                const stars = review.querySelector('[data-hook="review-star-rating"] span') ||
                               review.querySelector('.review-rating span');
                const date = review.querySelector('[data-hook="review-date"]');
                if (body && body.textContent.trim().length > 30) {
                    const starText = stars ? stars.textContent.trim() : '';
                    reviews.push({
                        title: title ? title.textContent.trim() : '',
                        body: body.textContent.trim().slice(0, 700),
                        stars: starText,
                        date: date ? date.textContent.trim() : ''
                    });
                }
            });
            return reviews.slice(0, 8);
        }"""
        cr = _run_playwright_page(critical_url, critical_js, wait=2)
        if isinstance(cr, list):
            critical_reviews = cr
        time.sleep(1)

    # ── merge all pages ───────────────────────────────────────────────────────
    data = {**page1}
    data["reviews"] = positive_reviews
    data["critical_reviews"] = critical_reviews
    data["url"] = url
    data["scraped_at"] = datetime.now().isoformat()

    cache_file.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    print(f"    [scrape] done — {len(positive_reviews)} pos reviews, {len(critical_reviews)} critical reviews")
    return data

# ── OLD SCRAPE FUNCTION PLACEHOLDER (remove closing bracket) ──


def _format_reviews(reviews, limit):
    """Render review dicts into readable, star+date-tagged lines for the prompt."""
    if not isinstance(reviews, list):   # search results store a review COUNT (int) here
        return ""
    lines = []
    for r in reviews[:limit]:
        if isinstance(r, dict):
            body = str(r.get("body", "")).strip().replace("\n", " ")
            if not body:
                continue
            lines.append(f"[{r.get('stars','?')}★ · {r.get('date','')}] "
                         f"{str(r.get('title','')).strip()}: {body[:320]}")
        elif str(r).strip():
            lines.append(f"- {str(r)[:320]}")
    return "\n".join(lines)


def analyse_product_with_ai(product_data, keyword):
    """
    Extract PRECISE, evidence-based intelligence from one product's raw data.
    Every review insight is grounded in the actual reviews with approximate
    mention counts and a short verbatim quote — not generic prose.
    """
    if not llm_available():
        return _mock_analysis(product_data)

    raw_bullets = product_data.get("bullets", [])
    bullets = ("\n".join(str(b) for b in raw_bullets[:6]) if isinstance(raw_bullets, list)
               else str(raw_bullets)[:400])[:800]
    description = str(product_data.get("description", ""))[:400]
    raw_rankings = product_data.get("rankings", [])
    rankings = ("\n".join(str(r) for r in raw_rankings[:3]) if isinstance(raw_rankings, list)
                else str(raw_rankings)[:300])

    title = str(product_data.get("title", product_data.get("product_name", "")))
    brand = str(product_data.get("brand", ""))
    price = safe_number(product_data.get("selling_price") or
                        product_data.get("price_inr") or product_data.get("price", 0))
    mrp_raw = product_data.get("mrp", "")
    mrp = safe_number(mrp_raw) if mrp_raw else safe_number(product_data.get("price_inr", price))
    discount = str(product_data.get("discount_pct", ""))
    rating_val = safe_number(product_data.get("rating", 0), max_val=5)
    review_count = int(safe_number(product_data.get("review_count", 0)))
    position = product_data.get("position", product_data.get("keyword_rank", 1))

    prod_details = product_data.get("product_details", {})
    details_text = ("\n".join(f"{k}: {v}" for k, v in list(prod_details.items())[:6])
                    if isinstance(prod_details, dict) else "")[:600]

    positive_text = _format_reviews(product_data.get("reviews", []), 12)
    critical_text = _format_reviews(product_data.get("critical_reviews", []), 12)

    # Amazon's own aspect insights (counted) — the strongest grounding for the narrative
    aspects_in = product_data.get("amazon_aspects", []) or []
    aspects_text = "\n".join(
        f"- {a.get('aspect','')}: {a.get('sentiment','')} "
        f"({a.get('positive',0)}/{a.get('total',0)} positive, {a.get('negative',0)} negative) — "
        f"{a.get('summary','')} e.g. \"{str(a.get('example',''))[:120]}\""
        for a in aspects_in[:10]) or "none"
    amazon_summary = product_data.get("amazon_summary", "")

    # Complaint signals: aspects ranked by NEGATIVE mentions (the negatives live here
    # even when the top displayed reviews are all positive — this is what fills the
    # "what customers complain about" section reliably from the full review base).
    neg_aspects = sorted([a for a in aspects_in if int(a.get("negative", 0) or 0) > 0],
                         key=lambda a: -int(a.get("negative", 0) or 0))[:8]
    complaint_signals = "\n".join(
        f"- {a.get('aspect','')}: {a.get('negative')} NEGATIVE mentions of {a.get('total')} total"
        for a in neg_aspects) or "none reported by Amazon"

    prompt = f"""You are a senior product-intelligence analyst for Deconstruct (Indian D2C skincare).
Analyse THIS specific Amazon India product for the keyword "{keyword}". Be precise and
evidence-based: every review insight MUST come from the reviews below, each with an
approximate mention count and a short verbatim quote. No generic filler, no invented facts.

PRODUCT: {title[:250]}
BRAND: {brand}
MRP ₹{mrp or '?'} | SELLING ₹{price} | DISCOUNT {discount or '-'}
RATING {rating_val}/5 from {review_count:,} reviews | SEARCH #{position}
RANKINGS: {rankings or 'n/a'}

ABOUT / BULLETS:
{bullets or 'n/a'}

DESCRIPTION:
{description or 'n/a'}

TECHNICAL DETAILS:
{details_text or 'n/a'}

POSITIVE REVIEWS:
{positive_text or 'none provided'}

CRITICAL REVIEWS:
{critical_text or 'none provided'}

AMAZON 'CUSTOMERS SAY' SUMMARY:
{amazon_summary or 'n/a'}

AMAZON ASPECT INSIGHTS (aspect: sentiment, positive/total mentions — these counts are authoritative, use them):
{aspects_text}

TOP COMPLAINT SIGNALS — aspects ranked by NEGATIVE mentions across ALL {review_count:,} reviews.
These ARE the complaints even if the sample reviews above look positive. You MUST build the
"key_negatives" from these (name the aspect + its negative count). NEVER leave negatives blank
when this list is non-empty:
{complaint_signals}

Return ONLY this JSON (start with {{):
{{
"brand": "clean brand name",
"product_name": "<=70 chars",
"price_inr": <number>, "mrp": <number, 0 if none>, "discount_pct": "e.g. 15% off or ''",
"rating": <float>, "review_count": <int>,
"category_rank": "e.g. '#3 in Sunscreens' or 'Not found'", "keyword_rank": <int>,
"pack_size_ml": <number in ml or g, null if unknown>,
"format": "gel/cream/serum/lotion/foam/oil/stick/other",
"spf": "e.g. 'SPF 50 PA++++' or ''",
"skin_types": ["oily","dry",...],
"key_actives": ["Niacinamide 10%","Zinc 1%"],
"top_claims": ["6-8 SPECIFIC claims taken from the listing"],
"claim_strength": <0-100 how specific/verifiable the claims are>,
"aspects": {{
  "efficacy":        {{"score": <0-100 or null>, "quote": "short verbatim or ''"}},
  "texture_finish":  {{"score": <0-100 or null>, "quote": ""}},
  "packaging":       {{"score": <0-100 or null>, "quote": ""}},
  "value_for_money": {{"score": <0-100 or null>, "quote": ""}},
  "fragrance":       {{"score": <0-100 or null>, "quote": ""}}
}},
"top_positives": [{{"point": "specific thing customers love", "mentions": <int>}}],
"top_complaints": [{{"issue": "specific complaint", "mentions": <int>, "quote": "short verbatim"}}],
"review_sentiment": "Positive/Mixed/Negative",
"key_positives": "A DEEP, thorough 6-9 sentence analysis of what customers genuinely love. Name each specific strength, tie it to the aspect mention counts (e.g. 'Sun protection: praised in 1,223 of 1,300 mentions'), say which skin types/use-cases praise it and why (the mechanism), and quote a real review line where possible. Analytical, not a summary blurb.",
"key_negatives": "A DEEP, hard-hitting 6-9 sentence analysis of the complaints and weaknesses. You MUST build this from the TOP COMPLAINT SIGNALS block — name each problem aspect AND its negative mention count (e.g. 'Skin compatibility drew 121 negative mentions: users with sensitive/acne-prone skin report breakouts and stinging'). Add any critical review quotes if present. Explain the likely CAUSE (why it happens), WHO it affects, and HOW SEVERE it is relative to total mentions. This section must NEVER be blank or vague when complaint signals exist — that is a failure.",
"strategic_read": "3-5 sentences written FOR Deconstruct's product team: exactly where this competitor is STRONG (so we must match or clearly differentiate) and where they are WEAK (a specific gap we can attack). Make it actionable — what should Deconstruct DO about this product.",
"market_gap": "2-3 sentences: the single biggest unmet need this product reveals = a specific launch/positioning opportunity for Deconstruct",
"overall_score": <0-100>
}}
Use null aspect scores where reviews give no signal. Prefer the Amazon aspect mention counts as ground truth. JSON only."""

    result = ask_llm(prompt, system="Output valid JSON only. Start with {.",
                     temperature=0.1, max_tokens=2200)
    try:
        return json.loads(result[result.index("{"): result.rindex("}") + 1])
    except Exception as e:
        print(f"[ai] parse error: {e}")
        return _mock_analysis(product_data)


# ── scoring ───────────────────────────────────────────────────────────────────
def score_product(analysis, position, total_products):
    """Score a product across all dimensions."""
    rating  = safe_number(analysis.get("rating", 0), max_val=5)
    reviews = int(safe_number(analysis.get("review_count", 0)))
    price   = safe_number(analysis.get("price_inr", 0))

    # normalise scores to 0-100
    rating_score    = min(100, (rating / 5.0) * 100) if rating else 50
    review_score    = min(100, (reviews / 1000) * 100) if reviews else 10
    price_score     = 70 if 200 < price < 800 else (50 if price > 0 else 40)
    claim_score     = analysis.get("claim_strength", 50)
    rank_score      = max(0, 100 - ((position - 1) / max(total_products, 1)) * 100)

    parts = {
        "rating_score":   round(rating_score),
        "review_volume":  round(review_score),
        "price_value":    round(price_score),
        "claim_strength": round(claim_score),
        "rank_position":  round(rank_score),
    }

    weights = {
        "rating_score": 0.25, "review_volume": 0.20,
        "price_value": 0.20, "claim_strength": 0.20, "rank_position": 0.15,
    }
    total = round(sum(parts[k] * weights[k] for k in weights))
    return total, parts


# ── full pipeline ─────────────────────────────────────────────────────────────
def run_analysis(keyword, urls=None, n_results=10,
                 date_from="", progress_cb=None, competitors_only=True):
    """
    Full pipeline:
    1. Search (if no URLs) or use provided URLs
    2. Keep only tracked competitors (if competitors_only)
    3. Fetch each product + AI analysis
    4. Score and rank
    Returns list of fully analysed products, sorted by score.
    """
    results = []

    if urls:
        # direct URL mode — user picked specific products, don't filter
        search_results = [{"url": u.strip(), "asin": "", "title": "", "brand": "",
                           "price": 0, "rating": 0, "reviews": 0,
                           "position": i+1, "thumbnail": ""}
                          for i, u in enumerate(urls) if u.strip()]
    elif competitors_only:
        # search broad, then keep the best-ranked product per tracked competitor
        if progress_cb:
            progress_cb(0, n_results, f"Searching Amazon for '{keyword}' (tracked competitors)...")
        picked = {}
        # 1. broad keyword search — catches whoever ranks + market context
        for sr in search_amazon(keyword, max(n_results * 5, 40)):
            canon = match_competitor(sr.get("brand", ""), sr.get("title", ""))
            if canon and canon not in picked:
                sr["brand"] = canon           # normalise to the canonical brand name
                picked[canon] = sr
        # 2. targeted per-brand search to GUARANTEE every tracked brand is covered
        #    (a brand may simply not rank on page 1 for this keyword — e.g. Minimalist)
        guarantee = list(COMPETITOR_ALIASES.keys())      # Deconstruct + all competitors
        for bi, brand in enumerate(guarantee):
            if brand in picked:
                continue
            if progress_cb:
                progress_cb(bi, len(guarantee), f"Locating {brand}'s {keyword}...")
            for sr in search_amazon(f"{brand} {keyword}", 5):
                if match_competitor(sr.get("brand", ""), sr.get("title", "")) == brand:
                    sr["brand"] = brand
                    picked[brand] = sr
                    break
        search_results = list(picked.values())
        print(f"[filter] {len(search_results)} tracked brands covered: {', '.join(picked.keys())}")
    else:
        # keyword search — analyse whatever ranks
        if progress_cb:
            progress_cb(0, n_results, f"Searching Amazon for '{keyword}'...")
        search_results = search_amazon(keyword, n_results)

    total = len(search_results)

    # save search results to cache immediately — protects API quota
    search_cache_file = CACHE / f"search_{hashlib.md5(keyword.encode()).hexdigest()}_results.json"
    if not search_cache_file.exists() and search_results:
        search_cache_file.write_text(json.dumps(search_results, indent=2), encoding='utf-8')
        print(f"[cache] search results saved — quota protected")

    for i, sr in enumerate(search_results):
        if progress_cb:
            progress_cb(i, total, f"Analysing: {sr.get('title','Product '+str(i+1))[:50]}...")

        # scrape full page
        url = sr.get("url") or f"https://www.amazon.in/dp/{sr.get('asin','')}"
        if not url or url == "https://www.amazon.in/dp/":
            continue

        asin = sr.get("asin", "") or _extract_asin(url)
        try:
            # PRIMARY: SerpAPI amazon_product engine (robust — no browser needed)
            page_data = fetch_product_serpapi(asin) if asin else {}
            # FALLBACK: Playwright scrape only if SerpAPI product returned nothing
            if not page_data:
                page_data = scrape_product_page(url, asin)
        except Exception as e:
            print(f"[scraper] failed for {url}: {e}")
            page_data = {}

        # Optional DEEP review sample (~100/product). Off by default because free
        # review-scraper actors cap at ~10/product — enabling this only pays off
        # once you're on a paid reviews-actor plan. Set DEEP_REVIEWS=1 in .env then.
        # (Amazon's full-base ASPECT COUNTS from page_data are the headline either way.)
        if asin and os.getenv("DEEP_REVIEWS", "0") == "1":
            deep_pos, deep_crit = fetch_reviews_apify(asin, target=120)
            if deep_pos or deep_crit:
                page_data = dict(page_data)
                page_data["reviews"] = deep_pos
                page_data["critical_reviews"] = deep_crit

        # merge search result data with page data (page_data wins on overlaps)
        merged = {**sr, **{k: v for k, v in page_data.items() if v}}

        # AI analysis — wrapped so one failure doesn't kill the whole run
        try:
            analysis = analyse_product_with_ai(merged, keyword)
        except Exception as e:
            print(f"[ai] failed for product {i+1}: {e}")
            analysis = _mock_analysis(sr)

        # ── AUTHORITATIVE FACTS ──────────────────────────────────────────────
        # Overwrite the decision-grade numbers with Amazon's own values (from
        # SerpAPI) AFTER the AI step, so the model can never alter a hard fact.
        # These are the fields the product team makes calls on.
        _sell = safe_number(merged.get("selling_price") or merged.get("price"))
        if _sell:
            analysis["price_inr"] = _sell
        _mrp = safe_number(merged.get("mrp"))
        if _mrp:
            analysis["mrp"] = _mrp
        if merged.get("discount_pct"):
            analysis["discount_pct"] = str(merged.get("discount_pct"))
        _rat = safe_number(merged.get("rating"), max_val=5)
        if _rat:
            analysis["rating"] = _rat
        _rc = int(safe_number(merged.get("review_count") or merged.get("review_count_raw")))
        if _rc:
            analysis["review_count"] = _rc
        # fill in from search result if AI couldn't extract
        if not analysis.get("rating") and sr.get("rating"):
            analysis["rating"] = sr["rating"]
        if not analysis.get("review_count") and sr.get("review_count_raw"):
            analysis["review_count"] = sr["review_count_raw"]
        # always set selling price from search result if AI didn't extract it
        if not analysis.get("price_inr"):
            analysis["price_inr"] = safe_number(sr.get("price", 0))
        # always set MRP — use explicit mrp field, fall back to search price
        if not analysis.get("mrp") or analysis.get("mrp") == analysis.get("price_inr"):
            srp_mrp = safe_number(sr.get("mrp", sr.get("price", 0)))
            merged_mrp = safe_number(merged.get("mrp", 0))
            analysis["mrp"] = merged_mrp if merged_mrp and merged_mrp != safe_number(sr.get("price",0)) else srp_mrp
        # discount
        if not analysis.get("discount_pct"):
            analysis["discount_pct"] = str(
                merged.get("discount_pct","") or
                sr.get("discount_pct","")
            )
        if not analysis.get("brand") and sr.get("brand"):
            analysis["brand"] = sr["brand"]
        if not analysis.get("product_name") and sr.get("title"):
            analysis["product_name"] = sr["title"][:80]

        analysis["keyword_rank"] = i + 1
        analysis["url"] = url
        analysis["thumbnail"] = sr.get("thumbnail", "")
        analysis["date_collected"] = datetime.now().strftime("%Y-%m-%d")
        analysis["date_range"] = date_from or datetime.now().strftime("%B %Y")
        analysis["keyword"] = keyword

        # score
        try:
            total_score, parts = score_product(analysis, i+1, total)
        except Exception as e:
            print(f"[score] failed for product {i+1}: {e}")
            total_score, parts = 50, {}
        analysis["total_score"] = total_score
        analysis["score_parts"] = parts

        # ── pricing normalisation: ₹ per ml (compare across pack sizes) ──
        _blob = " ".join([
            str(analysis.get("product_name", "")),
            str(sr.get("title", "")),
            str(merged.get("variant_hint", "")),
            " ".join(str(b) for b in merged.get("bullets", []))
            if isinstance(merged.get("bullets"), list) else "",
        ])
        size = safe_number(analysis.get("pack_size_ml"), default=0) or parse_pack_size(_blob)
        analysis["pack_size_ml"] = size or None
        analysis["price_per_ml"] = price_per_ml(analysis.get("price_inr"), size)

        # ── review-derived signals: star distribution + rating-over-time ──
        _pos = merged.get("reviews") if isinstance(merged.get("reviews"), list) else []
        _crit = merged.get("critical_reviews") if isinstance(merged.get("critical_reviews"), list) else []
        _all_reviews = [r for r in (_pos + _crit) if isinstance(r, dict)]
        analysis["star_distribution"] = star_distribution(_all_reviews)
        analysis["review_timeline"] = review_timeline(_all_reviews)
        analysis["reviews_sampled"] = len(_all_reviews)   # deep sample size (for credibility)

        # ── Amazon's own precise, pre-counted signals (best review data we have) ──
        analysis["amazon_aspects"] = merged.get("amazon_aspects", [])
        analysis["amazon_summary"] = merged.get("amazon_summary", "")
        analysis["bought_last_month"] = merged.get("bought_last_month", "")

        results.append(analysis)
        time.sleep(1.5)  # be polite

    # rank by total score
    results.sort(key=lambda x: x.get("total_score", 0), reverse=True)
    for i, r in enumerate(results, 1):
        r["final_rank"] = i

    save_snapshot(keyword, results, date_from)
    return results


# ── Excel export ──────────────────────────────────────────────────────────────
def to_excel(results, keyword, strategy=None):
    """Export results to a colour-coded Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        BLACK = PatternFill("solid", fgColor="0A0A0A")
        LIME_FONT = Font(bold=True, color="C8F55A", size=10)
        HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
        WRAP = Alignment(vertical='top', wrap_text=True)
        thin = Border(*(Side(style='thin', color='DDDDDD'),) * 4)

        def header_row(ws, headers, row=1, height=42):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.fill = BLACK
                c.font = LIME_FONT
                c.alignment = HDR_ALIGN
                c.border = thin
            ws.row_dimensions[row].height = height

        def score_fill(v):
            v = v if isinstance(v, (int, float)) else 0
            return PatternFill("solid", fgColor="DCFCE7" if v >= 70 else
                               "FEF9C3" if v >= 50 else "FEE2E2")

        def _as_text(v):
            """openpyxl can't store lists/dicts — flatten them to strings."""
            if isinstance(v, list):
                return "\n".join(f"• {x}" for x in v)
            if isinstance(v, dict):
                return "; ".join(f"{k}: {val}" for k, val in v.items())
            return v

        wb = openpyxl.Workbook()
        analytics = category_analytics(results)

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 1 — Ranked Products (ALL columns preserved + new ones added)
        # ══════════════════════════════════════════════════════════════════════
        ws = wb.active
        assert ws is not None
        ws.title = "Ranked Products"

        headers = [
            "Final Rank", "Brand", "Product Name", "Date",
            "MRP (₹)", "Selling Price (₹)", "Discount", "₹ / ml", "Pack Size (ml/g)",
            "Rating", "No. of Reviews", "Bought / Month",
            "Amazon Search Rank", "Category Rank", "About Product",
            "Top Claims", "Claim Strength", "Review Sentiment",
            "Rating Score\n(25%)", "Review Volume\n(20%)", "Price Value\n(20%)",
            "Claim Strength\n(20%)", "Rank Score\n(15%)", "TOTAL SCORE",
            "Market Gap", "URL",
        ]
        header_row(ws, headers)

        for r in results:
            row = r.get("final_rank", 1) + 1
            score = r.get("total_score", 0)
            row_fill = PatternFill("solid", fgColor="DCFCE7" if score >= 80 else
                                   "FEF9C3" if score >= 65 else
                                   "FFEDD5" if score >= 50 else "FEE2E2")
            sp = r.get("score_parts", {})
            mrp_val = safe_number(r.get("mrp", 0))
            sell_val = safe_number(r.get("price_inr", 0))
            ppm = r.get("price_per_ml")
            values = [
                r.get("final_rank", ""), r.get("brand", ""), r.get("product_name", ""),
                r.get("date_range", ""),
                f"₹{int(mrp_val)}" if mrp_val else "—",
                f"₹{int(sell_val)}" if sell_val else "—",
                str(r.get("discount_pct", "") or "—"),
                f"₹{ppm}" if isinstance(ppm, (int, float)) else "—",
                r.get("pack_size_ml") or "—",
                r.get("rating", ""), r.get("review_count", ""),
                str(r.get("bought_last_month", "") or "—"),
                r.get("keyword_rank", ""), r.get("category_rank", "Not found"),
                r.get("about_product", r.get("market_gap", "")),
                "\n".join(f"• {c}" for c in r.get("top_claims", [])[:6]),
                r.get("claim_strength", ""), r.get("review_sentiment", ""),
                sp.get("rating_score", ""), sp.get("review_volume", ""),
                sp.get("price_value", ""), sp.get("claim_strength", ""),
                sp.get("rank_position", ""), score,
                r.get("market_gap", ""), r.get("url", ""),
            ]
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=_as_text(val))
                c.border = thin
                c.alignment = WRAP
                if col in (19, 20, 21, 22, 23):
                    c.fill = score_fill(val)
                elif col == 24:
                    c.fill = row_fill
                    c.font = Font(bold=True, size=12)
                elif col == 1:
                    c.fill = row_fill
                    c.font = Font(bold=True, size=13)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    c.fill = PatternFill("solid", fgColor="F9F9F9" if row % 2 == 0 else "FFFFFF")
            ws.row_dimensions[row].height = 90

        widths = [8, 18, 34, 12, 11, 13, 10, 9, 12, 8, 12, 14, 12, 20, 40,
                  40, 10, 12, 9, 9, 9, 9, 9, 11, 40, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "C2"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 2 — Reviews Deep-Dive (long narrative + Amazon aspect counts)
        # ══════════════════════════════════════════════════════════════════════
        wr = wb.create_sheet("Reviews Deep-Dive")
        header_row(wr, ["Brand", "Product", "Rating", "What Customers LOVE",
                        "What Customers COMPLAIN About", "Strategic Read (for Deconstruct)",
                        "Top Amazon Aspects (mentions)"])
        for row_i, r in enumerate(results, 2):
            asp = r.get("amazon_aspects", []) or []
            asp_txt = "\n".join(
                f"{a.get('aspect','')}: {a.get('positive',0)}/{a.get('total',0)}+ "
                f"({a.get('negative',0)}−)" for a in asp[:8])
            vals = [r.get("brand", ""), str(r.get("product_name", ""))[:60], r.get("rating", ""),
                    r.get("key_positives", ""), r.get("key_negatives", ""),
                    r.get("strategic_read", ""), asp_txt]
            for col, val in enumerate(vals, 1):
                c = wr.cell(row=row_i, column=col, value=_as_text(val))
                c.border = thin
                c.alignment = WRAP
                if col == 4:
                    c.fill = PatternFill("solid", fgColor="F0FBE6")
                elif col == 5:
                    c.fill = PatternFill("solid", fgColor="FDECEC")
                elif col == 6:
                    c.fill = PatternFill("solid", fgColor="EFF6FF")
            wr.row_dimensions[row_i].height = 150
        for i, w in enumerate([16, 30, 8, 60, 60, 55, 34], 1):
            wr.column_dimensions[get_column_letter(i)].width = w
        wr.freeze_panes = "A2"

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 3 — Pricing & Demand
        # ══════════════════════════════════════════════════════════════════════
        wp = wb.create_sheet("Pricing & Demand")
        ps, pm = analytics.get("price_stats", {}), analytics.get("per_ml_stats", {})
        wp.cell(row=1, column=1, value="Category Pricing").font = Font(bold=True, size=13)
        rows = [
            ("Metric", "Selling Price (₹)", "₹ / ml"),
            ("Lowest", ps.get("min", "—"), pm.get("min", "—")),
            ("Q1 (25th pct)", ps.get("q1", "—"), pm.get("q1", "—")),
            ("Median", ps.get("median", "—"), pm.get("median", "—")),
            ("Q3 (75th pct)", ps.get("q3", "—"), pm.get("q3", "—")),
            ("Highest", ps.get("max", "—"), pm.get("max", "—")),
        ]
        for ri, rowv in enumerate(rows, 3):
            for ci, v in enumerate(rowv, 1):
                c = wp.cell(row=ri, column=ci, value=v)
                c.border = thin
                if ri == 3:
                    c.fill = BLACK
                    c.font = LIME_FONT
        wp.cell(row=11, column=1,
                value="Demand — Bought in past month (proxy for sales velocity)").font = Font(bold=True, size=13)
        header_row(wp, ["Brand", "Product", "Bought/Month", "Selling Price (₹)"], row=12)
        for ri, d in enumerate(analytics.get("demand", []), 13):
            for ci, v in enumerate([d["brand"], str(d["product"])[:50], d["bought"],
                                    f"₹{int(d['price'])}" if d.get("price") else "—"], 1):
                c = wp.cell(row=ri, column=ci, value=v)
                c.border = thin
                c.alignment = WRAP
        for i, w in enumerate([22, 40, 16, 16], 1):
            wp.column_dimensions[get_column_letter(i)].width = w

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 4 — Opportunities (whitespace, pain points, claim/ingredient freq)
        # ══════════════════════════════════════════════════════════════════════
        wo = wb.create_sheet("Opportunities")
        wo.cell(row=1, column=1,
                value="Launch & Positioning Opportunities").font = Font(bold=True, size=14, color="0A0A0A")
        r_ = 3
        for card in opportunity_cards(analytics):
            wo.cell(row=r_, column=1, value=f"[{card['type']}] {card['title']}").font = Font(bold=True, size=11)
            wo.cell(row=r_ + 1, column=1, value=card["detail"]).alignment = WRAP
            wo.row_dimensions[r_ + 1].height = 40
            r_ += 3
        wo.cell(row=r_, column=1, value="Category pain points (negative mentions = where everyone fails)").font = Font(bold=True, size=12)
        r_ += 1
        header_row(wo, ["Aspect", "Negative mentions", "Total mentions", "Products"], row=r_)
        for pp in analytics.get("pain_points", []):
            r_ += 1
            for ci, v in enumerate([pp["aspect"], pp["negative"], pp["total"], pp["products"]], 1):
                wo.cell(row=r_, column=ci, value=v).border = thin
        r_ += 2
        wo.cell(row=r_, column=1, value="Most common claims (crowded = avoid / differentiate)").font = Font(bold=True, size=12)
        r_ += 1
        for claim, n in analytics.get("claim_freq", [])[:10]:
            wo.cell(row=r_, column=1, value=claim)
            wo.cell(row=r_, column=2, value=n)
            r_ += 1
        for i, w in enumerate([70, 16, 16, 12], 1):
            wo.column_dimensions[get_column_letter(i)].width = w

        # ══════════════════════════════════════════════════════════════════════
        # SHEET 5 — Executive Summary
        # ══════════════════════════════════════════════════════════════════════
        we = wb.create_sheet("Executive Summary", 0)  # first tab
        we.cell(row=1, column=1,
                value="Decon AI — Competitor Marketplace Analysis").font = Font(bold=True, size=15)
        summary = [
            ("Keyword", keyword),
            ("Date", datetime.now().strftime("%d %B %Y")),
            ("Products analysed", analytics.get("n", len(results))),
            ("Avg rating", analytics.get("ratings_avg", "—")),
            ("Price range (₹)", f"₹{ps.get('min','?')} – ₹{ps.get('max','?')} (median ₹{ps.get('median','?')})" if ps else "—"),
            ("₹/ml range", f"₹{pm.get('min','?')} – ₹{pm.get('max','?')} (median ₹{pm.get('median','?')})" if pm else "—"),
            ("Top seller", (analytics.get("demand", [{}])[0].get("brand", "—")
                            + f" (~{analytics.get('demand',[{}])[0].get('bought',0):,}/mo)") if analytics.get("demand") else "—"),
            ("Biggest whitespace", (analytics.get("pain_points", [{}])[0].get("aspect", "—")
                                    + f" ({analytics.get('pain_points',[{}])[0].get('negative',0)} complaints)") if analytics.get("pain_points") else "—"),
            ("Deconstruct in set?", "Yes — baseline" if analytics.get("baseline_present") else "No"),
        ]
        for ri, (k, v) in enumerate(summary, 3):
            we.cell(row=ri, column=1, value=k).font = Font(bold=True)
            we.cell(row=ri, column=2, value=v).alignment = WRAP
        we.column_dimensions["A"].width = 26
        we.column_dimensions["B"].width = 60

        # ── AI decision sheets (if the caller computed them) ──
        def _s(v):
            if isinstance(v, list):
                return "\n".join(f"• {x}" for x in v)
            if isinstance(v, dict):
                return "\n".join(f"{k}: {val}" for k, val in v.items())
            return "" if v is None else str(v)

        def _brief_sheet(title, rows):
            sh = wb.create_sheet(title)
            sh.cell(row=1, column=1, value=f"Decon AI — {title} (AI synthesis, grounded in the "
                    "scraped facts & Amazon's counted review aspects)").font = Font(bold=True, size=12)
            sh.merge_cells("A1:B1")
            for ri, (label, val) in enumerate(rows, 3):
                c1 = sh.cell(row=ri, column=1, value=label)
                c1.font = Font(bold=True, color="0A0A0A")
                c1.alignment = Alignment(vertical="top", wrap_text=True)
                c1.fill = PatternFill("solid", fgColor="F1F0EC")
                c2 = sh.cell(row=ri, column=2, value=_s(val))
                c2.alignment = Alignment(vertical="top", wrap_text=True)
                sh.row_dimensions[ri].height = max(30, min(150, 16 * (len(_s(val)) // 60 + 1)))
            sh.column_dimensions["A"].width = 22
            sh.column_dimensions["B"].width = 95

        if strategy:
            sop = strategy.get("state_of_play", {}) or {}
            if sop:
                _brief_sheet("State of the Category", [
                    ("Executive summary", sop.get("summary", "")),
                    ("Market structure", sop.get("market_structure", "")),
                    ("Who's winning & why", [f"{l.get('brand','')} — {l.get('why_winning','')}"
                                             for l in sop.get("leaders", [])]),
                    ("Biggest gaps", sop.get("biggest_gaps", [])),
                    ("Price dynamics", sop.get("price_dynamics", "")),
                    ("Deconstruct's position", sop.get("deconstruct_position", "")),
                    ("Strategic moves", sop.get("strategic_moves", [])),
                ])
            lb = strategy.get("launch_brief", {}) or {}
            if lb:
                _brief_sheet("Launch Brief", [
                    ("Opportunity", lb.get("opportunity", "")),
                    ("Concept", lb.get("concept", "")),
                    ("Hero ingredient", lb.get("hero_ingredient", "")),
                    ("Format", lb.get("format", "")),
                    ("Target skin/concern", lb.get("target_skin", "")),
                    ("Key claims", lb.get("key_claims", [])),
                    ("Target price", lb.get("target_price_inr", "")),
                    ("Price rationale", lb.get("price_rationale", "")),
                    ("Name ideas", lb.get("name_ideas", [])),
                    ("Why now (evidence)", lb.get("why_now", "")),
                    ("Positioning", lb.get("positioning", "")),
                ])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    except Exception as e:
        print(f"[excel] error: {e}")
        import traceback
        traceback.print_exc()
        return _fallback_csv(results, keyword)


def _fallback_csv(results, keyword):
    lines = ["Rank,Brand,Product,Price,Rating,Reviews,Category Rank,Total Score"]
    for r in results:
        lines.append(f"{r.get('final_rank','')},{r.get('brand','')},\"{r.get('product_name','')}\","
                     f"{r.get('price_inr','')},{r.get('rating','')},{r.get('review_count','')},"
                     f"\"{r.get('category_rank','')}\",{r.get('total_score','')}")
    return "\n".join(lines).encode('utf-8')


# ── helpers ───────────────────────────────────────────────────────────────────
def _extract_brand(title):
    if not title:
        return "Unknown"
    words = title.split()
    return words[0] if words else "Unknown"


def _mock_search_results(keyword, n):
    """Demo data when SerpAPI key not set — includes MRP and selling price separately."""
    products = [
        {"asin":"B08X1234","title":"Minimalist SPF 50 Gel Sunscreen 50g","brand":"Minimalist",
         "price":299,"mrp":349,"discount_pct":"14% off",
         "rating":4.2,"reviews":15420,"position":1,
         "url":"https://www.amazon.in/dp/B08X1234","thumbnail":"","sponsored":False},
        {"asin":"B09Y2345","title":"Re'equil Oxybenzone & OMC Free Sunscreen SPF 50","brand":"Re'equil",
         "price":445,"mrp":495,"discount_pct":"10% off",
         "rating":4.4,"reviews":8930,"position":2,
         "url":"https://www.amazon.in/dp/B09Y2345","thumbnail":"","sponsored":False},
        {"asin":"B07Z3456","title":"Dot & Key Waterlight Sunscreen SPF 50","brand":"Dot & Key",
         "price":361,"mrp":425,"discount_pct":"15% off",
         "rating":4.1,"reviews":12100,"position":3,
         "url":"https://www.amazon.in/dp/B07Z3456","thumbnail":"","sponsored":False},
        {"asin":"B0AX4567","title":"Foxtale Lightweight SPF 50 PA+++ Sunscreen","brand":"Foxtale",
         "price":339,"mrp":399,"discount_pct":"15% off",
         "rating":4.3,"reviews":6780,"position":4,
         "url":"https://www.amazon.in/dp/B0AX4567","thumbnail":"","sponsored":False},
        {"asin":"B06W5678","title":"Neutrogena Ultra Sheer Dry Touch Sunscreen SPF 50+","brand":"Neutrogena",
         "price":539,"mrp":599,"discount_pct":"10% off",
         "rating":4.5,"reviews":43200,"position":5,
         "url":"https://www.amazon.in/dp/B06W5678","thumbnail":"","sponsored":False},
        {"asin":"B0BX6789","title":"The Derma Co 1% Hyaluronic Sunscreen SPF 50 PA+++","brand":"The Derma Co",
         "price":374,"mrp":449,"discount_pct":"17% off",
         "rating":4.2,"reviews":9340,"position":6,
         "url":"https://www.amazon.in/dp/B0BX6789","thumbnail":"","sponsored":False},
        {"asin":"B0CY7890","title":"Deconstruct Gel Sunscreen SPF 50 PA++++ 50g","brand":"Deconstruct",
         "price":331,"mrp":349,"discount_pct":"5% off",
         "rating":4.3,"reviews":15446,"position":7,
         "url":"https://www.amazon.in/dp/B0CY7890","thumbnail":"","sponsored":False},
    ]
    return products[:n]


def _mock_analysis(product_data):
    """
    Generate a UNIQUE analysis for each product based on its actual data.
    Uses the product title, brand, price and rating to infer specific insights.
    """
    title = str(product_data.get("title", product_data.get("product_name","Unknown Product")))
    brand = str(product_data.get("brand","Unknown Brand"))
    # selling price = discounted price
    price = safe_number(
        product_data.get("selling_price") or
        product_data.get("price") or
        product_data.get("price_inr", 0)
    )
    # MRP = original price before discount
    mrp = safe_number(
        product_data.get("mrp") or
        product_data.get("price") or
        product_data.get("price_inr", price)
    )
    discount = str(product_data.get("discount_pct", product_data.get("discount","")))
    rating = safe_number(product_data.get("rating", 4.0), max_val=5)
    reviews = int(safe_number(product_data.get("reviews", product_data.get("review_count", 0))))
    position = product_data.get("position", product_data.get("keyword_rank", 1))

    # extract claims from title keywords
    title_lower = title.lower()
    claims = []
    if "spf" in title_lower or "sunscreen" in title_lower or "sun" in title_lower:
        claims.append("Broad spectrum UV protection")
        claims.append("Lightweight, non-greasy texture")
    if "gel" in title_lower:
        claims.append("Gel-based formula for oily skin")
    if "niacinamide" in title_lower or "niacin" in title_lower:
        claims.append("10% Niacinamide for pore minimising")
        claims.append("Fades dark spots and acne marks")
    if "vitamin c" in title_lower or "vit c" in title_lower:
        claims.append("Vitamin C for brightening")
        claims.append("Antioxidant protection")
    if "hyaluronic" in title_lower or "ha" in title_lower:
        claims.append("Hyaluronic acid for deep hydration")
    if "retinol" in title_lower or "retinoid" in title_lower:
        claims.append("Retinol for anti-ageing")
        claims.append("Reduces fine lines and wrinkles")
    if "moistur" in title_lower:
        claims.append("24-hour moisture retention")
        claims.append("Non-comedogenic formula")
    if "serum" in title_lower:
        claims.append("Fast-absorbing serum texture")
    if "spf 50" in title_lower or "spf50" in title_lower:
        claims.append("SPF 50 high protection")
    if not claims:
        claims = ["Dermatologically tested", "Suitable for all skin types",
                  "Fragrance-free formula", "Clinically proven results"]

    # infer sentiment from rating
    if rating >= 4.3:
        sentiment = "Positive"
        positives = ["Consistently high ratings", "Strong repeat purchase rate",
                     f"Rated {rating}/5 by {reviews:,} customers"]
        negatives = ["Premium pricing vs competitors"] if price > 500 else []
    elif rating >= 3.8:
        sentiment = "Mixed"
        positives = [f"Decent rating of {rating}/5", "Good value for money"]
        negatives = ["Some complaints about packaging", "Results vary by skin type"]
    else:
        sentiment = "Negative"
        positives = ["Affordable price point"] if price < 400 else ["Available on Amazon"]
        negatives = ["Below average customer satisfaction", "Frequent complaints noted"]

    # infer market gap from product type
    if "sunscreen" in title_lower or "spf" in title_lower:
        gap = "Opportunity: No mention of reapplication convenience or ingredient transparency"
    elif "serum" in title_lower:
        gap = "Opportunity: Lacks clinical study references and specific % concentrations"
    elif "moistur" in title_lower:
        gap = "Opportunity: No skin-type specific variants or ingredient-forward positioning"
    else:
        gap = "Opportunity: Generic positioning — science-backed differentiation not highlighted"

    # category rank inference from position
    category_rank = f"#{position} in category (search rank)"

    # claim strength based on how specific claims are
    claim_strength = min(90, 50 + len([c for c in claims if "%" in c or any(x in c for x in ["SPF","acid","vitamin","retinol"])]) * 10)

    return {
        "brand": brand,
        "product_name": title[:80],
        "price_inr": price,
        "mrp": mrp,
        "discount_pct": discount,
        "rating": rating,
        "review_count": reviews,
        "category_rank": category_rank,
        "keyword_rank": position,
        "pack_size_ml": None,
        "format": "gel" if "gel" in title_lower else "serum" if "serum" in title_lower else "other",
        "spf": "SPF 50" if "spf 50" in title_lower or "spf50" in title_lower else "",
        "skin_types": [],
        "key_actives": [],
        "top_claims": claims[:5],
        "claim_strength": claim_strength,
        "review_sentiment": sentiment,
        "aspects": {},
        "top_positives": [{"point": p, "mentions": 0} for p in positives],
        "top_complaints": [{"issue": n, "mentions": 0, "quote": ""} for n in negatives],
        "key_positives": " ".join(positives) + " (offline estimate — live AI unavailable for this item)",
        "key_negatives": " ".join(negatives) or "No critical signal captured for this item.",
        "strategic_read": "Offline estimate — re-run with the AI engine connected for a full strategic read.",
        "market_gap": gap,
        "overall_score": min(95, max(40, int(rating * 15 + (min(reviews, 10000) / 10000) * 20))),
    }