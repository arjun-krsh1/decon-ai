"""
hr_processor.py — Direct resume processing for up to 20 resumes.
Reads PDFs → extracts text → scores with Groq → outputs Excel.
"""

import json
import time
import io
import sys
import pathlib
sys.path.insert(0, '.')
from dotenv import load_dotenv
load_dotenv()
from llm import ask_llm, llm_available

SCORING_DIMENSIONS = [
    {"key": "skills_match",         "label": "Skills Match",         "weight": 35},
    {"key": "experience_relevance", "label": "Experience Relevance", "weight": 25},
    {"key": "education_fit",        "label": "Education Fit",        "weight": 15},
    {"key": "cultural_indicators",  "label": "Cultural Indicators",  "weight": 15},
    {"key": "growth_trajectory",    "label": "Growth Trajectory",    "weight": 10},
]


def extract_pdf_text(file_bytes):
    """Extract text from PDF bytes."""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            return "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception:
        try:
            # rough fallback
            import re
            text = file_bytes.decode('latin-1', errors='ignore')
            chunks = re.findall(r'[A-Za-z][A-Za-z0-9 ,.\-/\n]{10,}', text)
            return "\n".join(chunks[:150])
        except Exception:
            return ""


def score_candidate(resume_text, jd_text, extra_criteria, candidate_name):
    """Score a single candidate against the JD."""
    if not llm_available():
        return _mock_score(candidate_name)

    prompt = f"""You are an expert HR recruiter. Score this candidate against the JD.

JD:
{jd_text[:2000]}

EXTRA CRITERIA: {extra_criteria or 'None'}

RESUME ({candidate_name}):
{resume_text[:2000]}

Return ONLY a JSON object:
{{
  "name": "candidate full name from resume or '{candidate_name}'",
  "current_role": "current or most recent job title",
  "current_company": "current or most recent company",
  "years_experience": "total years (number only)",
  "education": "highest degree + field",
  "skills_match": 0-100,
  "experience_relevance": 0-100,
  "education_fit": 0-100,
  "cultural_indicators": 0-100,
  "growth_trajectory": 0-100,
  "strengths": ["strength 1", "strength 2", "strength 3"],
  "gaps": ["gap 1", "gap 2"],
  "notes": "any flags or standout qualities",
  "recommendation": "STRONG YES or YES or MAYBE or NO"
}}

JSON only. Start with {{"""

    result = ask_llm(prompt, system="You are a precise HR evaluator. Output JSON only.", temperature=0.1)
    try:
        start = result.index('{')
        end = result.rindex('}') + 1
        data = json.loads(result[start:end])
        # calculate total score
        total = (
            data.get("skills_match", 0) * 0.35 +
            data.get("experience_relevance", 0) * 0.25 +
            data.get("education_fit", 0) * 0.15 +
            data.get("cultural_indicators", 0) * 0.15 +
            data.get("growth_trajectory", 0) * 0.10
        )
        data["total_score"] = round(total)
        data["file_name"] = candidate_name
        return data
    except Exception as e:
        print(f"[hr] parse error for {candidate_name}: {e}")
        return _mock_score(candidate_name)


def _mock_score(name):
    return {
        "name": name, "current_role": "Unknown", "current_company": "Unknown",
        "years_experience": "?", "education": "Unknown",
        "skills_match": 60, "experience_relevance": 60, "education_fit": 60,
        "cultural_indicators": 60, "growth_trajectory": 60, "total_score": 60,
        "strengths": ["Could not parse resume"], "gaps": ["Resume parsing failed"],
        "notes": "Manual review needed", "recommendation": "MAYBE",
        "file_name": name
    }


def process_resumes(resume_files, jd_text, extra_criteria, progress_callback=None):
    """
    Process multiple resume files.
    resume_files: list of (filename, bytes) tuples
    Returns: list of scored candidate dicts, sorted by total_score
    """
    results = []
    total = len(resume_files)

    for i, (filename, file_bytes) in enumerate(resume_files):
        if progress_callback:
            progress_callback(i, total, filename)

        # extract text
        resume_text = extract_pdf_text(file_bytes)
        if not resume_text.strip():
            print(f"  [{filename}] could not extract text")
            results.append(_mock_score(filename))
            continue

        # score with AI
        candidate_name = pathlib.Path(filename).stem
        scored = score_candidate(resume_text, jd_text, extra_criteria, candidate_name)
        results.append(scored)

        # rate limit protection — 4 sec between calls
        if i < total - 1:
            time.sleep(4)

    # sort by total score
    results.sort(key=lambda x: x.get("total_score", 0), reverse=True)

    # add rank
    for i, r in enumerate(results, 1):
        r["rank"] = i

    return results


def results_to_excel(results):
    """Convert scored results to Excel bytes."""
    try:
        import openpyxl
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                      Border, Side)
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidate Rankings"

        # colours
        black = PatternFill("solid", fgColor="0A0A0A")
        lime   = PatternFill("solid", fgColor="C8F55A")
        green  = PatternFill("solid", fgColor="DCFCE7")
        yellow = PatternFill("solid", fgColor="FEF9C3")
        orange = PatternFill("solid", fgColor="FFEDD5")
        red    = PatternFill("solid", fgColor="FEE2E2")
        grey   = PatternFill("solid", fgColor="F5F5F5")

        thin = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        # header row
        headers = [
            "Rank", "Name", "Current Role", "Current Company",
            "Years Exp", "Education",
            "Skills Match\n(35%)", "Experience\n(25%)", "Education Fit\n(15%)",
            "Cultural\n(15%)", "Growth\n(10%)",
            "TOTAL\nSCORE",
            "Top Strengths", "Key Gaps",
            "Recommendation", "Notes", "File"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = black
            cell.font = Font(bold=True, color="C8F55A", size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                        wrap_text=True)
            cell.border = thin

        ws.row_dimensions[1].height = 40

        # data rows
        for r in results:
            row = r.get("rank", 1) + 1
            rec = r.get("recommendation", "MAYBE")
            total = r.get("total_score", 0)

            # row background based on recommendation
            row_fill = (green if rec == "STRONG YES" else
                        yellow if rec == "YES" else
                        orange if rec == "MAYBE" else red)

            values = [
                r.get("rank", ""),
                r.get("name", ""),
                r.get("current_role", ""),
                r.get("current_company", ""),
                r.get("years_experience", ""),
                r.get("education", ""),
                r.get("skills_match", ""),
                r.get("experience_relevance", ""),
                r.get("education_fit", ""),
                r.get("cultural_indicators", ""),
                r.get("growth_trajectory", ""),
                total,
                "\n".join(f"• {s}" for s in r.get("strengths", [])),
                "\n".join(f"• {g}" for g in r.get("gaps", [])),
                rec,
                r.get("notes", ""),
                r.get("file_name", ""),
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # score columns — colour by value
                if col in (7, 8, 9, 10, 11):
                    score = val if isinstance(val, (int, float)) else 0
                    if score >= 80:
                        cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    elif score >= 60:
                        cell.fill = PatternFill("solid", fgColor="FEF9C3")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")
                elif col == 12:  # total score
                    cell.font = Font(bold=True, size=12)
                    if total >= 85:
                        cell.fill = PatternFill("solid", fgColor="DCFCE7")
                        cell.font = Font(bold=True, size=12, color="166534")
                    elif total >= 70:
                        cell.fill = PatternFill("solid", fgColor="FEF9C3")
                    elif total >= 55:
                        cell.fill = PatternFill("solid", fgColor="FFEDD5")
                    else:
                        cell.fill = PatternFill("solid", fgColor="FEE2E2")
                elif col == 15:  # recommendation
                    cell.font = Font(bold=True)
                    cell.fill = row_fill
                elif col == 1:  # rank
                    cell.font = Font(bold=True, size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill = grey if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

            ws.row_dimensions[row].height = 60

        # column widths
        col_widths = [6, 22, 22, 22, 8, 20, 10, 10, 10, 10, 10, 10, 35, 30, 14, 30, 20]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # freeze header row
        ws.freeze_panes = "A2"

        # summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Decon AI — ATS Evaluation Summary"
        ws2["A1"].font = Font(bold=True, size=14)

        strong_yes = sum(1 for r in results if r.get("recommendation") == "STRONG YES")
        yes = sum(1 for r in results if r.get("recommendation") == "YES")
        maybe = sum(1 for r in results if r.get("recommendation") == "MAYBE")
        no = sum(1 for r in results if r.get("recommendation") == "NO")

        summary_data = [
            ("", ""),
            ("Total Candidates Evaluated", len(results)),
            ("", ""),
            ("STRONG YES (85-100)", strong_yes),
            ("YES (70-84)", yes),
            ("MAYBE (55-69)", maybe),
            ("NO (0-54)", no),
            ("", ""),
            ("Top Candidate", results[0].get("name", "") if results else ""),
            ("Top Score", results[0].get("total_score", "") if results else ""),
        ]

        for row_idx, (label, value) in enumerate(summary_data, 2):
            ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=bool(label))
            ws2.cell(row=row_idx, column=2, value=value)

        ws2.column_dimensions["A"].width = 35
        ws2.column_dimensions["B"].width = 25

        # save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    except ImportError:
        # fallback: CSV if openpyxl not installed
        return _results_to_csv(results)


def _results_to_csv(results):
    """CSV fallback if openpyxl not available."""
    lines = ["Rank,Name,Current Role,Company,Years Exp,Total Score,Recommendation,Strengths,Gaps,Notes"]
    for r in results:
        strengths = "; ".join(r.get("strengths", []))
        gaps = "; ".join(r.get("gaps", []))
        lines.append(
            f"{r.get('rank','')},{r.get('name','')},{r.get('current_role','')},"
            f"{r.get('current_company','')},{r.get('years_experience','')},"
            f"{r.get('total_score','')},{r.get('recommendation','')},"
            f"\"{strengths}\",\"{gaps}\",\"{r.get('notes','')}\"")
    return "\n".join(lines).encode('utf-8')