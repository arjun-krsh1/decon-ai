"""
excel_report.py — turns the search_intel insights into a brand-manager workbook.

Not a data dump: every sheet opens with a header block that states WHAT it shows,
HOW it's calculated, HOW ACCURATE it is, and a plain-English glossary of every
column. Numbers are formatted, colour-coded and charted so a brand manager gets
the answer at a glance and can justify it upward.

Pure openpyxl (no new dependency). build_workbook() returns .xlsx bytes.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule

# ── Deconstruct palette ──
INK, LIME, CREAM, SALMON, WHITE = "0A0A0A", "C8F55A", "F6F5F2", "F4A99A", "FFFFFF"
GREY, ZEBRA = "5A5850", "FAFAF8"
GOOD, WARN, CRIT, NEUT = "2E7D46", "B7791F", "C1443A", "5A5850"

_thin = Side(style="thin", color="D8D6CE")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)


GLOSSARY = {
    "Impressions": "How many times a Deconstruct link was shown in Google Search results.",
    "Clicks": "How many times people clicked through to the site.",
    "CTR %": "Click-through rate = Clicks ÷ Impressions × 100. Higher = the listing earns its views.",
    "CTR": "Click-through rate = Clicks ÷ Impressions × 100.",
    "Exp %": "Expected CTR for that ranking position (industry curve) — the benchmark we compare against.",
    "Position": "Average Google ranking position (1 = very top of results; lower is better).",
    "Avg Position": "Impression-weighted average ranking position (1 = top; lower is better).",
    "Pos": "Average Google ranking position (1 = top).",
    "Queries": "Number of distinct search terms grouped here.",
    "Category": "Product category inferred from the query and verified against the page URL.",
    "Keyword": "The exact search term a person typed into Google.",
    "Intent": "Why they searched: product / commercial / navigational / reputation / etc.",
    "Class": "How the query was classified — product, brand, commercial, noise, etc.",
    "Impr %": "This group's share of total impressions.",
    "Page": "The Deconstruct URL Google showed.",
    "Verdict": "Current-state read of the category (Dominant / Scale up / Emerging / Underperforming).",
    "Clicks/day": "Average clicks per day that month.",
    "Impr/day": "Average impressions per day that month.",
    "Attribute A": "A product attribute (format / active / concern) found in the query.",
    "Attribute B": "A second attribute found in the same query.",
    "Country": "Searcher's country.", "Device": "Device used to search.",
}

VERDICT_COLOR = {"Dominant": GOOD, "Scale up": WARN, "Emerging": NEUT, "Underperforming": CRIT}


def _fmt(label):
    h = label.lower()
    if "%" in h or "ctr" in h or "exp" in h:
        return "0.00"
    if "pos" in h:
        return "0.0"
    if any(k in h for k in ["impr", "click", "quer", "day", "volume"]):
        return "#,##0"
    return None


def _title_block(ws, title, what, how, accuracy):
    for col in range(1, 9):
        ws.cell(1, col).fill = _fill(INK)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c = ws.cell(1, 1, title)
    c.font = Font(bold=True, size=15, color=LIME)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 30
    r = 2
    for label, val in [("WHAT THIS SHOWS", what), ("HOW IT'S CALCULATED", how), ("HOW ACCURATE", accuracy)]:
        for col in range(1, 9):
            ws.cell(r, col).fill = _fill(CREAM)
        lc = ws.cell(r, 1, label)
        lc.font = Font(bold=True, size=9, color=GREY)
        lc.alignment = Alignment(vertical="top", horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        vc = ws.cell(r, 2, val)
        vc.font = Font(size=10, color=INK)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30
        r += 1
    return r + 1


def _glossary(ws, start_row, columns):
    items = [(lbl, GLOSSARY[lbl]) for (_, lbl, _) in columns if lbl in GLOSSARY]
    if not items:
        return start_row
    c = ws.cell(start_row, 1, "WHAT EACH COLUMN MEANS")
    c.font = Font(bold=True, size=9, color=GREY)
    r = start_row + 1
    for term, mean in items:
        tc = ws.cell(r, 1, term)
        tc.font = Font(bold=True, size=9, color=INK)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        mc = ws.cell(r, 2, mean)
        mc.font = Font(size=9, color=GREY)
        mc.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        r += 1
    return r + 1


def _table(ws, start_row, columns, rows, scale_key=None, verdict_key=None):
    hr = start_row
    for i, (key, label, w) in enumerate(columns, 1):
        c = ws.cell(hr, i, label)
        c.font = Font(bold=True, size=10, color=WHITE)
        c.fill = _fill(INK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[hr].height = 24
    r = hr + 1
    for row in rows:
        for i, (key, label, w) in enumerate(columns, 1):
            v = row.get(key, "")
            c = ws.cell(r, i, v)
            c.border = BORDER
            c.font = Font(size=10, color=INK)
            f = _fmt(label)
            if f and isinstance(v, (int, float)):
                c.number_format = f
            c.alignment = Alignment(horizontal="left" if i == 1 else "center",
                                    vertical="center", wrap_text=(i == 1))
            if (r - hr) % 2 == 0:
                c.fill = _fill(ZEBRA)
            if verdict_key and key == verdict_key and isinstance(v, str):
                c.font = Font(size=10, bold=True, color=VERDICT_COLOR.get(v, INK))
        r += 1
    ws.freeze_panes = ws.cell(hr + 1, 1)
    last = r - 1
    if scale_key and last > hr:
        idx = next((i for i, (k, l, w) in enumerate(columns, 1) if k == scale_key), None)
        if idx:
            col = get_column_letter(idx)
            ws.conditional_formatting.add(
                f"{col}{hr + 1}:{col}{last}",
                ColorScaleRule(start_type="min", start_color=WHITE, end_type="max", end_color=LIME))
    return hr, last


def _data_sheet(wb, name, title, what, how, accuracy, columns, rows,
                scale_key=None, verdict_key=None, glossary=True):
    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    nr = _title_block(ws, title, what, how, accuracy)
    if glossary:
        nr = _glossary(ws, nr, columns)
    _table(ws, nr, columns, rows or [{}], scale_key=scale_key, verdict_key=verdict_key)
    return ws


def _note(ws, row, text, color=GREY, bold=False, size=10, fill=None):
    for col in range(1, 9):
        if fill:
            ws.cell(row, col).fill = _fill(fill)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row, 1, text)
    c.font = Font(size=size, color=color, bold=bold)
    c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws.row_dimensions[row].height = max(18, 15 * (len(text) // 95 + 1))
    return row + 1


# ── special sheets ───────────────────────────────────────────────────────────
def _start_here(wb, ins, tab_index):
    ws = wb.create_sheet("Start Here")
    ws.sheet_view.showGridLines = False
    m = ins["meta"]
    r = _title_block(
        ws, "Deconstruct — Search Intelligence Report",
        f"Your Google Search Console demand for {m.get('date_range','')}, cleaned of brand-name "
        "noise and turned into insights a brand manager can act on and justify.",
        "Read directly from your Search Console export (Google's own recorded numbers — nothing "
        "estimated), then classified with an auditable skincare lexicon so real demand is separated from noise.",
        f"{100 - next((c['impr_pct'] for c in ins['classes'] if c['class']=='ambiguous'), 0):.1f}% of "
        "impressions are classified into a named group; every figure traces to a rule shown on its tab.")
    r = _note(ws, r, "TERMS — what the numbers mean", color=INK, bold=True, size=11)
    for term in ["Impressions", "Clicks", "CTR %", "Position", "Queries", "Category", "Intent"]:
        r = _note(ws, r, f"• {term} — {GLOSSARY[term]}")
    r += 1
    r = _note(ws, r, "WHAT'S IN THIS WORKBOOK", color=INK, bold=True, size=11)
    for tab, desc in tab_index:
        rr = ws.cell(r, 1, tab)
        rr.font = Font(bold=True, size=10, color=INK)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        d = ws.cell(r, 2, desc)
        d.font = Font(size=10, color=GREY)
        d.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        r += 1
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 13
    return ws


def _exec_summary(wb, es):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    r = _title_block(
        ws, "Executive Summary",
        "The one-glance read: what Deconstruct is in search, what's working, what's leaking, and what to do first.",
        "AI synthesis (Groq/Llama) written strictly on top of the computed numbers in this workbook — it interprets, it does not invent figures.",
        "Grounded in the same source numbers as every other tab; each point cites its metric.")
    r = _note(ws, r, es.get("headline", "Connect Groq for the AI summary."), color=INK, bold=True, size=12,
              fill=CREAM)
    r += 1
    for title_, key, color in [("✅ WHAT'S WORKING", "working", GOOD),
                               ("⚠️ WHAT'S NOT WORKING", "not_working", CRIT),
                               ("🎯 PRIORITIES (do these first)", "priorities", INK)]:
        r = _note(ws, r, title_, color=color, bold=True, size=11)
        for item in es.get(key, []) or ["—"]:
            r = _note(ws, r, f"•  {item}", color=INK)
        r += 1
    if es.get("seasonality"):
        r = _note(ws, r, "📅 SEASONALITY", color=INK, bold=True, size=11)
        r = _note(ws, r, es["seasonality"], color=INK)
    ws.column_dimensions["A"].width = 22
    return ws


def _dashboard(wb, ins):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    _title_block(ws, "Dashboard — the picture at a glance",
                 "Category demand, momentum over time and search intent as charts.",
                 "Bars/lines are drawn directly from the Category Demand, Seasonality and Intent tabs — same numbers, visualised.",
                 "Exactly the tab figures; nothing new is introduced here.")
    # category source (rows 7+)
    cats = ins["categories"][:8]
    base = 7
    ws.cell(base, 1, "Category").font = Font(bold=True, color=WHITE)
    ws.cell(base, 2, "Impressions").font = Font(bold=True, color=WHITE)
    ws.cell(base, 1).fill = _fill(INK)
    ws.cell(base, 2).fill = _fill(INK)
    for i, c in enumerate(cats, 1):
        ws.cell(base + i, 1, c["category"])
        ws.cell(base + i, 2, c["impressions"]).number_format = "#,##0"
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Category demand (impressions)"
    bar.height, bar.width = 8, 13
    data = Reference(ws, min_col=2, min_row=base, max_row=base + len(cats))
    cat_ref = Reference(ws, min_col=1, min_row=base + 1, max_row=base + len(cats))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cat_ref)
    bar.legend = None
    ws.add_chart(bar, "D7")

    months = ins["trend"]["months"]
    mb = base + len(cats) + 3
    for j, h in enumerate(["Month", "Clicks/day", "Impr/day"], 1):
        cell = ws.cell(mb, j, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = _fill(INK)
    for i, mm in enumerate(months, 1):
        ws.cell(mb + i, 1, mm["month"])
        ws.cell(mb + i, 2, mm["clicks_per_day"]).number_format = "#,##0"
        ws.cell(mb + i, 3, mm["impr_per_day"]).number_format = "#,##0"
    line = LineChart()
    line.title = "Momentum — clicks & impressions per day"
    line.height, line.width = 8, 13
    ldata = Reference(ws, min_col=2, max_col=3, min_row=mb, max_row=mb + len(months))
    lcat = Reference(ws, min_col=1, min_row=mb + 1, max_row=mb + len(months))
    line.add_data(ldata, titles_from_data=True)
    line.set_categories(lcat)
    ws.add_chart(line, "D22")

    intents = ins["intents"][:6]
    ib = mb + len(months) + 3
    for j, h in enumerate(["Intent", "Impressions"], 1):
        cell = ws.cell(ib, j, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = _fill(INK)
    for i, it in enumerate(intents, 1):
        ws.cell(ib + i, 1, it["intent"])
        ws.cell(ib + i, 2, it["impressions"]).number_format = "#,##0"
    ibar = BarChart()
    ibar.type = "bar"
    ibar.title = "Search intent (impressions)"
    ibar.height, ibar.width = 7, 13
    idata = Reference(ws, min_col=2, min_row=ib, max_row=ib + len(intents))
    icat = Reference(ws, min_col=1, min_row=ib + 1, max_row=ib + len(intents))
    ibar.add_data(idata, titles_from_data=True)
    ibar.set_categories(icat)
    ibar.legend = None
    ws.add_chart(ibar, "D37")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    return ws


def build_workbook(ins, es, reads, working, notw):
    es = es or {}
    reads = reads or {}
    m = ins["meta"]
    wb = Workbook()
    wb.remove(wb.active)

    tab_index = [
        ("Executive Summary", "The one-glance read — what's working, what's not, what to do."),
        ("Dashboard", "Charts: category demand, momentum, intent."),
        ("Product Scorecard", "Every category graded Dominant / Scale up / Emerging / Underperforming."),
        ("Keyword Performance", "The keywords driving demand, with an AI read of what wins."),
        ("What's Working", "Winning categories, pages and brand pull — with the number."),
        ("What's Not Working", "Leaks and gaps, each with the fix."),
        ("Category Demand", "Full demand + CTR + position by category."),
        ("Seasonality", "Month, weekday and spike patterns over the window."),
        ("Demand Interlinks", "Which attributes people search together."),
        ("Opportunities", "Striking-distance and low-CTR quick wins."),
        ("Page Performance", "Every page: demand vs capture."),
        ("All Queries Classified", "The full audit — every query, its class and metrics."),
        ("Excluded Noise", "Exactly what was removed as non-skincare brand noise."),
        ("Reach", "Countries, devices and how you appear in search."),
        ("Methodology & Accuracy", "Every rule and every limitation, in the open."),
    ]
    _start_here(wb, ins, tab_index)
    _exec_summary(wb, es)
    _dashboard(wb, ins)

    # Product Scorecard (merge AI verdicts)
    readmap = {r.get("category", ""): r for r in reads.get("category_reads", [])}
    sc_rows = []
    for c in ins["categories"]:
        rd = readmap.get(c["category"], {})
        sc_rows.append({"Category": c["category"], "Impressions": c["impressions"], "Clicks": c["clicks"],
                        "CTR %": c.get("ctr_pct", 0), "Avg Position": c.get("avg_position", 0),
                        "Verdict": rd.get("label", ""), "What it means": rd.get("note", "")})
    _data_sheet(wb, "Product Scorecard", "Product Scorecard — where each category stands",
                "Every product category graded on current demand and how well it converts — so you know where to double down.",
                "Impressions/Clicks summed per category; CTR = clicks÷impressions; Avg Position is impression-weighted. "
                "Verdict is an AI read of demand size vs capture (current-state, not growth-over-time).",
                "Category totals are exact; the Verdict is an interpretation of those exact numbers.",
                [("Category", "Category", 20), ("Impressions", "Impressions", 14), ("Clicks", "Clicks", 11),
                 ("CTR %", "CTR %", 9), ("Avg Position", "Avg Position", 12), ("Verdict", "Verdict", 15),
                 ("What it means", "What it means", 60)],
                sc_rows, scale_key="Impressions", verdict_key="Verdict")

    # Keyword Performance (dedicated)
    kw_rows = [r for r in ins["query_table"] if r["class"] in ("product", "commercial")]
    kw_rows = sorted(kw_rows, key=lambda r: -r["impressions"])[:60]
    kw_cols = [("query", "Keyword", 40), ("category", "Category", 18), ("intent", "Intent", 14),
               ("impressions", "Impressions", 13), ("clicks", "Clicks", 10),
               ("ctr_pct", "CTR %", 9), ("position", "Position", 10)]
    kws = _data_sheet(wb, "Keyword Performance", "Keyword Performance — what people actually type",
                      "The individual search terms driving Deconstruct's demand, ranked by impressions, with category and intent.",
                      "Product/commercial queries only (brand-name noise removed); CTR = clicks÷impressions; Position is the Google average.",
                      "These are Google's recorded per-query numbers; categories are cross-checked against page URLs.",
                      kw_cols, [{**r, "ctr_pct": r["ctr_pct"]} for r in kw_rows],
                      scale_key="impressions")
    if reads.get("keyword_read"):
        # append the AI keyword read below the table
        last = kws.max_row + 2
        _note(kws, last, "AI KEYWORD READ", color=INK, bold=True, size=11)
        _note(kws, last + 1, reads["keyword_read"], color=INK, fill=CREAM)

    _data_sheet(wb, "What's Working", "What's Working — keep doing this",
                "The categories, pages and brand pull that are already winning.",
                "Winning categories are the top by clicks; winning pages rank in the top 10 with healthy CTR; brand pull is navigational click share.",
                "Straight from the source numbers; 'why' is a plain-language label, not a new metric.",
                [("Area", "Area", 12), ("Item", "Item", 46), ("Clicks", "Clicks", 11),
                 ("Impressions", "Impressions", 13), ("CTR %", "CTR %", 9), ("Position", "Position", 10),
                 ("Why it works", "Why it works", 42)],
                working, glossary=False)

    _data_sheet(wb, "What's Not Working", "What's Not Working — fix these",
                "Where demand exists but the brand isn't capturing it — the fastest wins.",
                "Leaking pages/queries rank in the top 10 but earn below 60% of the expected CTR for their position; "
                "striking-distance items rank 8–20 (one push from page 1).",
                "Each row is a real page/query with its real metrics; the action is the recommended next step.",
                [("Issue", "Issue", 34), ("Item", "Item", 44), ("Impressions", "Impressions", 13),
                 ("CTR %", "CTR %", 9), ("Position", "Position", 10), ("What to do", "What to do", 48)],
                notw, glossary=False)

    _data_sheet(wb, "Category Demand", "Category Demand — size of each opportunity",
                "How much search demand each product category pulls, and how well it converts.",
                "Impressions/Clicks summed over the product queries in each category; CTR = clicks÷impressions; Avg Position impression-weighted.",
                "Exact category totals from the classified queries (brand-name noise removed).",
                [("category", "Category", 22), ("queries", "Queries", 10), ("impressions", "Impressions", 14),
                 ("clicks", "Clicks", 11), ("ctr_pct", "CTR %", 9), ("avg_position", "Avg Position", 12)],
                ins["categories"], scale_key="impressions")

    months = ins["trend"]["months"]
    _data_sheet(wb, "Seasonality", "Seasonality — how demand moves over time",
                "Search demand by month, by weekday, and the biggest single days in the window.",
                "Averaged from the daily totals in your export. NOTE: this tab is TOTAL traffic per day — per-category "
                "seasonal curves need category-segmented exports (see Methodology).",
                "Daily totals are Google's exact recorded numbers; monthly figures are their averages.",
                [("month", "Month", 14), ("days", "Days", 8), ("clicks_per_day", "Clicks/day", 13),
                 ("impr_per_day", "Impr/day", 13), ("avg_position", "Avg Position", 12)],
                months, scale_key="clicks_per_day")

    _data_sheet(wb, "Demand Interlinks", "Demand Interlinks — what's searched together",
                "Which product attributes (format, active, concern) show up together in the same searches — the shape of demand.",
                "For every product query we tag its attributes, then count attribute pairs weighted by that query's impressions.",
                "Derived from the exact query text; weights are real impression sums.",
                [("a", "Attribute A", 22), ("b", "Attribute B", 22), ("impressions", "Impressions", 14)],
                ins["interlinks"], scale_key="impressions")

    opp = [{"Type": "Striking distance (rank 8-20)", "Query/Page": x["query"], "Impressions": x["impressions"],
            "CTR %": "", "Position": x["position"]} for x in ins["striking"][:12]]
    opp += [{"Type": "Low CTR (ranks, under-clicked)", "Query/Page": x["query"], "Impressions": x["impressions"],
             "CTR %": x["ctr"], "Position": x["position"]} for x in ins["low_ctr"][:12]]
    _data_sheet(wb, "Opportunities", "Opportunities — the quickest wins",
                "Queries that are one small step from a lot more traffic.",
                "Striking-distance = average position 8–20 (nearly page 1). Low-CTR = ranks in the top 10 but earns below "
                "60% of the expected CTR for its position — usually a title/snippet rewrite.",
                "Real per-query metrics from your export; the expected-CTR benchmark is a standard position curve.",
                [("Type", "Type", 30), ("Query/Page", "Keyword", 40), ("Impressions", "Impressions", 14),
                 ("CTR %", "CTR %", 9), ("Position", "Position", 10)],
                opp, scale_key="Impressions")

    _data_sheet(wb, "Page Performance", "Page Performance — demand vs capture by URL",
                "Every top page: how much it's shown, how well it converts, and where it ranks.",
                "Canonical page URLs from the export; CTR = clicks÷impressions; category read from the URL path.",
                "Google's exact per-page numbers — the URL removes any ambiguity about what the page is.",
                [("page", "Page", 52), ("category", "Category", 18), ("impressions", "Impressions", 13),
                 ("clicks", "Clicks", 10), ("ctr_pct", "CTR %", 9), ("position", "Position", 10)],
                ins["page_table"][:60], scale_key="impressions")

    _data_sheet(wb, "All Queries Classified", "All Queries — the full audit",
                "Every query in the export with the class we assigned it, its category and intent, and its metrics.",
                "Each query classified by the lexicon (first matching rule wins); category verified against page URLs.",
                "This is the complete, auditable backing for every other tab — nothing is hidden.",
                [("query", "Keyword", 44), ("class", "Class", 12), ("category", "Category", 18),
                 ("intent", "Intent", 16), ("impressions", "Impressions", 13), ("clicks", "Clicks", 10),
                 ("ctr_pct", "CTR %", 9), ("position", "Position", 10)],
                ins["query_table"], scale_key="impressions")

    _data_sheet(wb, "Excluded Noise", "Excluded Noise — what we removed and why",
                "The branded queries that are NOT about skincare (careers, the dictionary sense, etc.), removed so they don't inflate demand.",
                "Flagged by the noise lexicon (internship, career, meaning, definition, …) when no skincare term is present.",
                "This is the transparency behind the accuracy claim — you can see exactly what was set aside.",
                [("query", "Keyword", 44), ("intent", "Intent", 16), ("impressions", "Impressions", 13),
                 ("clicks", "Clicks", 10), ("ctr_pct", "CTR %", 9), ("position", "Position", 10)],
                [r for r in ins["query_table"] if r["class"] == "noise"])

    reach = [{"Dimension": "Country", "Value": g["label"], "Clicks": g["clicks"],
              "Impressions": g["impressions"], "CTR %": g["ctr"]} for g in ins["geo"]]
    reach += [{"Dimension": "Device", "Value": d["label"], "Clicks": d["clicks"],
               "Impressions": d["impressions"], "CTR %": d["ctr"]} for d in ins["device"]]
    reach += [{"Dimension": "Search appearance", "Value": a["label"], "Clicks": a["clicks"],
               "Impressions": a["impressions"], "CTR %": a["ctr"]} for a in ins["appearance"]]
    _data_sheet(wb, "Reach", "Reach — who sees you and how",
                "Where the demand comes from: countries, devices and how Deconstruct appears in results.",
                "Straight from the Countries, Devices and Search-appearance tabs of your export.",
                "Google's exact recorded numbers per dimension.",
                [("Dimension", "Dimension", 18), ("Value", "Value", 24), ("Clicks", "Clicks", 11),
                 ("Impressions", "Impressions", 14), ("CTR %", "CTR %", 9)],
                reach)

    lx = ins["lexicon_sizes"]
    meth_ws = wb.create_sheet("Methodology & Accuracy")
    meth_ws.sheet_view.showGridLines = False
    r = _title_block(meth_ws, "Methodology & Accuracy — nothing is a black box",
                     "Exactly how every number in this workbook is produced, and the limits of this data.",
                     "Read this to justify any figure to a brand manager or auditor.",
                     f"{100 - next((c['impr_pct'] for c in ins['classes'] if c['class']=='ambiguous'), 0):.1f}% of impressions classified; noise isolated and listed.")
    r = _note(meth_ws, r, "HOW THE NUMBERS ARE MADE", color=INK, bold=True, size=11)
    for t in [
        f"Source: your Search Console export ({m.get('date_range','')}) — Queries, Pages, daily Chart, Countries, Devices, Search-appearance. Google's own numbers; nothing estimated.",
        f"Export filter: {m.get('query_filter','')} — every query contains the brand word, so raw counts mix real demand with brand-name noise.",
        f"Classification: rule-based lexicon ({lx['niche']} skincare terms, {lx['commercial']} commercial, {lx['reputation']} reputation, {lx['noise']} noise, {lx['tool']} tool). First matching rule wins.",
        "Categories: from query text, cross-checked against canonical page URLs (/collections/, /products/).",
        "Opportunity rule: CTR compared to a position-based expected-CTR curve; flagged below 60% of expected while ranking in the top 10.",
    ]:
        r = _note(meth_ws, r, "•  " + t)
    r += 1
    r = _note(meth_ws, r, "WHAT THIS DATA CAN'T SHOW (yet)", color=CRIT, bold=True, size=11)
    for t in [
        "Top-1,000 cap: the export holds the top 1,000 queries; the long tail beyond isn't included.",
        "Brand filter: filtered to +deconstruct, so pure non-branded demand ('best sunscreen for oily skin') isn't captured — the live API can pull it.",
        "Per-category seasonality: the daily tab is totals only, so 'is sunscreen growing vs winter' needs category-segmented exports (two periods) which merge into monthly curves per category.",
    ]:
        r = _note(meth_ws, r, "•  " + t)
    meth_ws.column_dimensions["A"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
