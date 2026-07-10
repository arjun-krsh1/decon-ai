"""
report.py — beautified, self-explanatory Excel for Comp's Nemesis.

Reuses the proven styling engine in excel_report.py: every tab opens with
WHAT THIS SHOWS · HOW IT'S CALCULATED · HOW ACCURATE · a glossary of every
column, colour-coded, with a charted Dashboard and a Start-Here guide.
Every row carries its public product URL for verifiability.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font

import excel_report as X   # generic styling helpers (title block, glossary, table, charts)

# ── Comp's-Nemesis-specific glossary (merged into the shared one) ──
_TERMS = {
    "Brand": "The competitor. Deconstruct is the baseline everything is compared against.",
    "Product": "Product title exactly as listed on the brand's own website.",
    "₹": "Current selling price on the brand's website (₹).",
    "Old ₹": "Selling price at the previous snapshot (₹).",
    "New ₹": "Selling price at the current snapshot (₹).",
    "Δ%": "Change in selling price between snapshots (negative = a price cut).",
    "MRP": "Maximum retail / compare-at price (₹) — the struck-through price.",
    "Discount %": "Discount off MRP = (MRP − price) ÷ MRP × 100.",
    "Available": "In stock right now, from the store's own stock flag.",
    "Type": "Product category/type as the brand labels it.",
    "Published": "Date the product page went live on the brand's site.",
    "Source": "The public product URL — open it to verify any figure on the row.",
    "Event": "Stock change detected between snapshots (went out of stock / restocked).",
    "Products": "Number of live products in the brand's catalog.",
    "Avg ₹": "Average selling price across the brand's real products (junk SKUs excluded).",
    "On discount %": "Share of the brand's products currently discounted.",
    "Avg discount %": "Average discount depth across the discounted products.",
    "Out of stock": "Number of products currently unavailable.",
    "Newest launch": "Publish date of the brand's most recent product.",
    "Old discount": "Discount % at the previous snapshot.",
    "New discount": "Discount % at the current snapshot.",
}

_TAB_INDEX = [
    ("Dashboard", "Charts — catalog size, discounting and launches per brand at a glance."),
    ("Catalog Benchmark", "Every brand vs Deconstruct: size, price, discounting, stock."),
    ("New Launches", "Products that appeared since the last snapshot (early warning)."),
    ("Price Moves", "Products whose selling price changed between snapshots."),
    ("Discount Moves", "Products whose discount depth changed (campaign / pressure)."),
    ("Stock Flips", "Products that went out of stock or restocked."),
    ("Removals", "Products that disappeared (discontinued / delisted)."),
    ("Full Catalog", "Every tracked product with price, MRP, discount, stock and its URL."),
    ("Methodology", "Exactly how each number is produced, and the data source."),
]


def _flat(diffs, key, keys):
    out = []
    for _b, d in (diffs or {}).items():
        for x in d.get(key, []):
            if key == "price_changes" and x.get("noise"):
                continue
            out.append({k: x.get(k, "") for k in keys})
    return out


def _start_here(wb, curr, prev_date, curr_date):
    ws = wb.create_sheet("Start Here")
    ws.sheet_view.showGridLines = False
    total = sum(len(v) for v in (curr or {}).values())
    r = X._title_block(
        ws, "Comp's Nemesis — Competitor Intelligence",
        f"What {len(curr or {})} D2C skincare brands are doing on their own websites — launches, prices, "
        "discounts and stock — tracked over time and benchmarked against Deconstruct.",
        "Collected automatically every 2 days from each brand's PUBLIC Shopify storefront JSON "
        "(/products.json — the same data their site serves). No logins, no grey-hat. Comparisons come "
        "from diffing consecutive snapshots.",
        f"100% first-party public data — {total:,} products across {len(curr or {})} brands. Every row "
        "carries its public product URL; open it to verify any number.")
    r = X._note(ws, r, f"Snapshot: {curr_date}   ·   compared against: {prev_date or '(baseline — no prior yet)'}",
                color=X.INK, bold=True, size=11, fill=X.CREAM)
    r += 1
    r = X._note(ws, r, "TERMS — what the columns mean", color=X.INK, bold=True, size=11)
    for term in ["Product", "₹", "MRP", "Discount %", "Available", "Δ%", "Event", "Newest launch", "Source"]:
        r = X._note(ws, r, f"•  {term} — {X.GLOSSARY.get(term, '')}")
    r += 1
    r = X._note(ws, r, "WHAT'S IN THIS WORKBOOK", color=X.INK, bold=True, size=11)
    for tab, desc in _TAB_INDEX:
        cell = ws.cell(r, 1, tab)
        cell.font = Font(bold=True, size=10, color=X.INK)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        d = ws.cell(r, 2, desc)
        d.font = Font(size=10, color=X.GREY)
        r += 1
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 13


def _dashboard(wb, bench, diffs):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    X._title_block(ws, "Dashboard — the picture at a glance",
                   "Catalog size, discount intensity and new-launch counts per brand.",
                   "Bars are drawn from the Catalog Benchmark and New Launches tabs — same numbers, visualised.",
                   "Exactly the tab figures; nothing new introduced here.")
    top = [b for b in bench][:11]
    base = 7
    ws.cell(base, 1, "Brand").font = Font(bold=True, color=X.WHITE)
    ws.cell(base, 1).fill = X._fill(X.INK)
    for j, h in enumerate(["Products", "On discount %", "New launches"], 2):
        cc = ws.cell(base, j, h)
        cc.font = Font(bold=True, color=X.WHITE)
        cc.fill = X._fill(X.INK)
    launch_by_brand = {b: len(d.get("launches", [])) for b, d in (diffs or {}).items()}
    for i, row in enumerate(top, 1):
        ws.cell(base + i, 1, row["brand"])
        ws.cell(base + i, 2, row["products"])
        ws.cell(base + i, 3, row["on_discount_%"])
        ws.cell(base + i, 4, launch_by_brand.get(row["brand"], 0))
    n = len(top)
    cats = Reference(ws, min_col=1, min_row=base + 1, max_row=base + n)

    bar1 = BarChart(); bar1.type = "bar"; bar1.title = "Catalog size (products) by brand"
    bar1.height, bar1.width = 9, 13; bar1.legend = None
    bar1.add_data(Reference(ws, min_col=2, min_row=base, max_row=base + n), titles_from_data=True)
    bar1.set_categories(cats)
    ws.add_chart(bar1, "F7")

    bar2 = BarChart(); bar2.type = "bar"; bar2.title = "% of catalog on discount"
    bar2.height, bar2.width = 9, 13; bar2.legend = None
    bar2.add_data(Reference(ws, min_col=3, min_row=base, max_row=base + n), titles_from_data=True)
    bar2.set_categories(cats)
    ws.add_chart(bar2, "F26")

    if any(launch_by_brand.values()):
        bar3 = BarChart(); bar3.type = "bar"; bar3.title = "New launches since last snapshot"
        bar3.height, bar3.width = 9, 13; bar3.legend = None
        bar3.add_data(Reference(ws, min_col=4, min_row=base, max_row=base + n), titles_from_data=True)
        bar3.set_categories(cats)
        ws.add_chart(bar3, "F45")
    ws.column_dimensions["A"].width = 20


def build_excel(curr: dict, diffs: dict, bench: list, prev_date, curr_date) -> bytes:
    X.GLOSSARY.update(_TERMS)
    wb = Workbook()
    _default = wb.active
    if _default is not None:
        wb.remove(_default)

    _start_here(wb, curr, prev_date, curr_date)
    _dashboard(wb, bench, diffs)

    X._data_sheet(wb, "Catalog Benchmark", "Catalog Benchmark — every brand vs Deconstruct",
                  "Each competitor's catalog size, pricing, discounting and stock, side by side. "
                  "Deconstruct is the top (baseline) row.",
                  "Aggregated from each brand's live public catalog; junk SKUs (₹1 promos, testers) "
                  "excluded from the price/discount stats.",
                  "Direct from each brand's public Shopify JSON — nothing estimated. Verify via any Source URL.",
                  [("brand", "Brand", 20), ("products", "Products", 11), ("avg_price", "Avg ₹", 10),
                   ("on_discount_%", "On discount %", 14), ("avg_discount_%", "Avg discount %", 14),
                   ("out_of_stock", "Out of stock", 12), ("newest_launch", "Newest launch", 14)],
                  bench, scale_key="products")

    X._data_sheet(wb, "New Launches", "New Launches — products that just appeared",
                  f"Products in the {curr_date} snapshot that were absent last time — often live "
                  "on the website before any public announcement.",
                  "Diff of product handles between the two most recent snapshots, per brand.",
                  "Public Shopify catalog; every row's Source is the live product page.",
                  [("brand", "Brand", 18), ("title", "Product", 44), ("price", "₹", 9),
                   ("product_type", "Type", 20), ("published_at", "Published", 12), ("url", "Source", 52)],
                  _flat(diffs, "launches", ["brand", "title", "price", "product_type", "published_at", "url"]))

    X._data_sheet(wb, "Price Moves", "Price Moves — who changed selling price",
                  "Same product, different selling price vs the previous snapshot. A cut on marketplace-facing "
                  "SKUs can signal rank defence or inventory pressure.",
                  "Compares each product's price across the two most recent snapshots (junk SKUs excluded).",
                  "Public Shopify price fields; open Source to confirm.",
                  [("brand", "Brand", 18), ("title", "Product", 40), ("old_price", "Old ₹", 10),
                   ("new_price", "New ₹", 10), ("delta_pct", "Δ%", 9), ("url", "Source", 52)],
                  _flat(diffs, "price_changes", ["brand", "title", "old_price", "new_price", "delta_pct", "url"]))

    X._data_sheet(wb, "Discount Moves", "Discount Moves — campaign & pressure signals",
                  "Products whose discount depth changed. Deep new discounts can flag excess stock or a push.",
                  "Compares each product's discount % (MRP vs price) across snapshots.",
                  "Public Shopify compare-at vs price fields.",
                  [("brand", "Brand", 18), ("title", "Product", 40), ("old_discount", "Old discount", 13),
                   ("new_discount", "New discount", 13), ("url", "Source", 52)],
                  _flat(diffs, "discount_changes", ["brand", "title", "old_discount", "new_discount", "url"]))

    X._data_sheet(wb, "Stock Flips", "Stock Flips — out-of-stock & restock events",
                  "Products that flipped availability since last snapshot. Repeated stock-outs = a hit they "
                  "can't keep up with, or a supply problem.",
                  "Compares each product's in-stock flag across the two most recent snapshots.",
                  "Public Shopify 'available' flag.",
                  [("brand", "Brand", 18), ("title", "Product", 42), ("event", "Event", 22), ("url", "Source", 52)],
                  _flat(diffs, "stock_changes", ["brand", "title", "event", "url"]))

    X._data_sheet(wb, "Removals", "Removals — products that disappeared",
                  "Product pages present last snapshot but gone now — discontinued or delisted.",
                  "Handles in the previous snapshot missing from the current one.",
                  "Public Shopify catalog diff.",
                  [("brand", "Brand", 18), ("title", "Product", 44), ("price", "₹", 9), ("url", "Source", 52)],
                  _flat(diffs, "removals", ["brand", "title", "price", "url"]))

    # cap per brand so the styled workbook stays memory-safe on a small host
    catalog = [{"brand": p["brand"], "title": p["title"], "product_type": p["product_type"],
                "price": p["price"], "mrp": p["mrp"], "discount_pct": p["discount_pct"],
                "available": p["available"], "published_at": p["published_at"], "url": p["url"]}
               for prods in (curr or {}).values() for p in prods[:150]]
    X._data_sheet(wb, "Full Catalog", "Full Catalog — tracked products",
                  "The monitored catalog across all brands — the raw evidence behind every other tab "
                  "(up to 150 products per brand; the complete catalog always lives in the database).",
                  "Live products from each brand's public /products.json.",
                  "First-party public data; each row's Source is its public URL.",
                  [("brand", "Brand", 18), ("title", "Product", 42), ("product_type", "Type", 18),
                   ("price", "₹", 9), ("mrp", "MRP", 9), ("discount_pct", "Discount %", 11),
                   ("available", "Available", 10), ("published_at", "Published", 12), ("url", "Source", 50)],
                  catalog, glossary=True)

    mws = wb.create_sheet("Methodology")
    mws.sheet_view.showGridLines = False
    r = X._title_block(mws, "Methodology & Data-Source Transparency",
                       "Exactly how every number in this workbook is produced.",
                       "Read this to justify any figure to a colleague or partner.",
                       "100% public first-party data; every row links to its public source.")
    for t in [
        "Source — each brand's own PUBLIC Shopify storefront JSON (/products.json): the same data their "
        "website renders. No logins, no scraping behind auth, no grey-hat methods.",
        "Traceability — every row's 'Source' column is the public product URL; open it to verify any figure.",
        "Launches / Removals — product handles that appear / disappear between two consecutive snapshots.",
        "Price / Discount moves — the same product's selling price or discount depth (MRP vs price) changing "
        "between snapshots.",
        "Stock flips — the Shopify 'available' flag changing between snapshots.",
        "Cadence — a GitHub Action collects a fresh snapshot automatically every 2 days; comparisons appear "
        "from the second snapshot onward.",
        "Junk SKUs — ₹1 promos, gift cards and testers are flagged out of price/discount stats but kept for "
        "launch detection (a new kit is still a real signal).",
        "Baseline — Deconstruct; every competitor is measured against it.",
    ]:
        r = X._note(mws, r, "•  " + t)
    mws.column_dimensions["A"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
