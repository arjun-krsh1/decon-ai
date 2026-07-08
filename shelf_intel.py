"""
shelf_intel.py — analytics for the Market Share / Digital-Shelf module.

Turns the raw marketplace listings into the outcomes the team acts on:
  * Share-of-Shelf   — % of top listings each brand holds (per platform + blended)
  * Share-of-Search  — ranking dominance (top-3 / top-10 presence, avg rank)
  * Assortment       — SKU count per brand per platform
  * Distribution gaps— brand × platform presence matrix
  * Demand proxies   — total review volume + avg rating per brand
  * Price tiers      — budget/mid/premium positioning per brand
  * Keyword analysis — what the top-ranked listings put in their titles, and an
                       AI read of WHY those products rank (themes/claims)
  * Deconstruct scorecard vs the tracked competitors

All the numbers are deterministic counts/maths over observed listings. Only the
"why do they rank" narrative uses the LLM (grounded in the scraped titles).
"""

from __future__ import annotations

import re
import json
import math
import statistics
from collections import Counter, defaultdict

import pathlib
from datetime import datetime

from llm import groq_chat, groq_available

BASELINE = "Deconstruct"
_SNAP = pathlib.Path("scraper/cache/shelf/snapshots")
_SNAP.mkdir(parents=True, exist_ok=True)


def save_shelf_snapshot(report):
    """Persist a run (scorecard + shelf share) for movement-over-time tracking."""
    try:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        slim = {"generated_at": datetime.now().isoformat(timespec="seconds"),
                "keywords": report.get("keywords", []),
                "scorecard": report.get("scorecard", []),
                "share_of_shelf": report.get("share_of_shelf", [])[:15]}
        (_SNAP / f"run_{stamp}.json").write_text(json.dumps(slim, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"[shelf] snapshot save failed: {e}")


def load_shelf_snapshots():
    runs = []
    for f in sorted(_SNAP.glob("run_*.json"), reverse=True):
        try:
            runs.append(json.loads(f.read_text(encoding="utf-8")))
        except Exception:
            continue
    return runs

# generic filler + non-differentiating words to ignore when mining title keywords.
# NOTE: we deliberately KEEP descriptive terms (gel, matte, cooling, brightening,
# niacinamide, watermelon...) — those are the "peculiar" keywords the team wants.
_STOP = set("""the a an and or for with of to in on at from by is are was your you this that
these those it its our my we all also more most top up new best buy online india indian
results result use uses using apply pump bottle tube ml g gm gram grams kg litre pack
combo set kit oz size ct count value pcs pc x rs inr price mrp sale deal offer flat off
type men women unisex pa spf skincare care product products formula formulation range
based free based-formula per each get amp nbsp skin face facial ideal suitable types
ml. gms qty""".split()) | {str(n) for n in range(0, 1001)}

# multi-word skincare claims/ingredients worth capturing as single precise tokens
_PHRASES = [
    "vitamin c", "hyaluronic acid", "salicylic acid", "glycolic acid", "lactic acid",
    "alpha arbutin", "kojic acid", "azelaic acid", "tranexamic acid", "ascorbic acid",
    "dark spot", "dark spots", "dark circle", "dark circles", "white cast", "no white cast",
    "dry touch", "oil free", "oil control", "matte finish", "water resistant", "water based",
    "broad spectrum", "rice water", "green tea", "aloe vera", "niacinamide", "retinol",
    "ceramide", "peptide", "collagen", "spf 50", "spf 30", "spf 40", "pa++++", "pa+++",
    "anti ageing", "anti aging", "anti pigmentation", "pore minimising", "brightening",
    "hydrating", "hydration", "cooling", "gel based", "lightweight", "non greasy",
    "acne prone", "sensitive skin", "oily skin", "dry skin", "de tan", "sun protection",
]


def _cat_words(category):
    """The category's own query words — obvious, so we drop them from its keyword list."""
    return {w for w in re.findall(r"[a-z0-9%+]+", str(category).lower()) if len(w) > 1}


def _price_tier(p):
    if not p:
        return None
    return "budget" if p < 350 else "mid" if p < 700 else "premium"


def _brand_of(listing):
    return listing.get("brand", "Unknown")


def share_of_shelf(listings, platform=None):
    """% of listings held by each brand (optionally within one platform)."""
    rows = [l for l in listings if platform is None or l["platform"] == platform]
    total = len(rows) or 1
    c = Counter(_brand_of(l) for l in rows)
    return [{"brand": b, "skus": n, "pct": round(100 * n / total, 1)}
            for b, n in c.most_common()]


def search_dominance(listings):
    """Per brand: top-3 and top-10 appearances and average rank (share of search)."""
    agg = defaultdict(lambda: {"top3": 0, "top10": 0, "ranks": []})
    for l in listings:
        r = l.get("rank", 99)
        a = agg[_brand_of(l)]
        a["ranks"].append(r)
        if r <= 3:
            a["top3"] += 1
        if r <= 10:
            a["top10"] += 1
    out = []
    for b, a in agg.items():
        out.append({"brand": b, "top3": a["top3"], "top10": a["top10"],
                    "avg_rank": round(statistics.mean(a["ranks"]), 1) if a["ranks"] else None,
                    "listings": len(a["ranks"])})
    out.sort(key=lambda x: (-x["top10"], x["avg_rank"] if x["avg_rank"] else 99))
    return out


def assortment_matrix(listings):
    """{brand: {platform: sku_count}} — breadth of range per platform."""
    m = defaultdict(lambda: defaultdict(int))
    for l in listings:
        m[_brand_of(l)][l["platform"]] += 1
    return {b: dict(pv) for b, pv in m.items()}


def demand_and_price(listings):
    """Per brand: total reviews (demand proxy), avg rating, avg price, price tier."""
    agg = defaultdict(lambda: {"reviews": 0, "ratings": [], "prices": []})
    for l in listings:
        a = agg[_brand_of(l)]
        a["reviews"] += l.get("reviews", 0)
        if l.get("rating"):
            a["ratings"].append(l["rating"])
        if l.get("price"):
            a["prices"].append(l["price"])
    out = []
    for b, a in agg.items():
        avg_price = round(statistics.mean(a["prices"])) if a["prices"] else 0
        out.append({"brand": b, "total_reviews": a["reviews"],
                    "avg_rating": round(statistics.mean(a["ratings"]), 2) if a["ratings"] else None,
                    "avg_price": avg_price, "price_tier": _price_tier(avg_price)})
    out.sort(key=lambda x: -x["total_reviews"])
    return out


# ── keyword analysis (distinctive, not generic) ───────────────────────────────
def _tokens(title, category="", exclude=frozenset()):
    """
    Precise tokens from a title: unigrams + curated skincare phrases, minus generic
    filler, the category's own query words (so 'sunscreen' is dropped under the
    sunscreen category), and brand-name words (a brand's name isn't a 'keyword').
    """
    t = " " + str(title).lower() + " "
    found = set()
    for ph in _PHRASES:                       # capture multi-word claims first
        if f" {ph} " in t or ph in t:
            found.add(ph)
    for w in re.findall(r"[a-z][a-z0-9%+]{2,}", t):
        if w not in _STOP:
            found.add(w)
    drop = _cat_words(category) | set(exclude)
    return {w for w in found if w not in drop}


def _brand_stop(listings):
    """Every word appearing in any brand name — excluded from keyword mining."""
    stop = set()
    for l in listings:
        for w in re.findall(r"[a-z0-9]+", str(l.get("brand", "")).lower()):
            if len(w) > 1:
                stop.add(w)
    return stop


def category_keywords(listings, top_rank=15, categories=None, exclude=frozenset()):
    """
    DISTINCTIVE keywords per category (TF-IDF style): terms frequent in a category's
    top listings but NOT generic across all categories. Answers 'under sunscreen, what
    words actually differentiate the winners' — gel, matte, brightening, not 'sunscreen'.
    Returns {category: [{"kw":, "listings":, "score":}]}.
    """
    cats = categories or sorted({l["keyword"] for l in listings})
    n_cats = max(len(cats), 1)
    cat_counter = defaultdict(Counter)      # category -> token -> #listings using it
    token_cats = defaultdict(set)           # token -> set of categories it appears in
    for l in listings:
        if l.get("rank", 99) > top_rank:
            continue
        cat = l["keyword"]
        for tok in _tokens(l["title"], cat, exclude):
            cat_counter[cat][tok] += 1
            token_cats[tok].add(cat)
    out = {}
    for cat in cats:
        scored = []
        for tok, cnt in cat_counter.get(cat, {}).items():
            idf = math.log((n_cats + 1) / len(token_cats[tok])) + 1   # rare-across-cats => high
            scored.append({"kw": tok, "listings": cnt, "score": round(cnt * idf, 1)})
        scored.sort(key=lambda x: (-x["score"], -x["listings"]))
        out[cat] = scored[:12]
    return out


def competitor_keywords(listings, brands, exclude=frozenset()):
    """
    For each tracked brand: the distinctive descriptors across ITS listings — e.g.
    'Dot & Key -> cooling, watermelon, hydration'. {brand: [(kw, count), ...]}.
    """
    agg = defaultdict(Counter)
    for l in listings:
        b = _brand_of(l)
        if b in brands:
            for tok in _tokens(l["title"], l["keyword"], exclude):
                agg[b][tok] += 1
    return {b: agg[b].most_common(12) for b in brands if agg.get(b)}


def product_keywords(listings, top_rank=5, per_platform=3, exclude=frozenset()):
    """
    Product-wise keyword bifurcation for the standout listings — e.g.
    'Dot & Key Watermelon Sunscreen -> watermelon, cooling, hydrating'.
    Returns a list of {platform, category, brand, title, url, keywords}.
    """
    rows = []
    seen = defaultdict(int)
    for l in sorted(listings, key=lambda x: (x["platform"], x["keyword"], x.get("rank", 99))):
        k = (l["platform"], l["keyword"])
        if l.get("rank", 99) <= top_rank and seen[k] < per_platform:
            toks = sorted(_tokens(l["title"], l["keyword"], exclude))
            if toks:
                rows.append({"platform": l["platform"], "category": l["keyword"],
                             "brand": l["brand"], "title": l["title"], "url": l.get("url", ""),
                             "keywords": toks[:8], "rank": l.get("rank")})
                seen[k] += 1
    return rows


def gap_analysis(cat_keywords, comp_keywords, listings, baseline="Deconstruct", exclude=frozenset()):
    """
    AI read of what the winners claim that Deconstruct does NOT — a precise,
    category-level 'what should we add to our titles/claims' brief.
    """
    if not groq_available():
        return {"per_category": {}, "actions": [], "summary": "Connect Groq for the gap analysis."}
    dec_tokens = sorted({t for l in listings if _brand_of(l) == baseline
                         for t in _tokens(l["title"], l["keyword"], exclude)})
    cat_lines = []
    for cat, kws in cat_keywords.items():
        cat_lines.append(f"{cat}: winning keywords = " + ", ".join(k["kw"] for k in kws[:10]))
    comp_lines = [f"{b}: " + ", ".join(k for k, _ in kws[:8]) for b, kws in comp_keywords.items()]
    prompt = f"""You are a D2C skincare category strategist for {baseline}.

WINNING KEYWORDS PER CATEGORY (distinctive terms the top-ranked listings use):
{chr(10).join(cat_lines)[:2500]}

WHAT EACH COMPETITOR EMPHASISES (their distinctive keywords):
{chr(10).join(comp_lines)[:2000]}

{baseline}'S OWN KEYWORDS (what our listings currently say):
{', '.join(dec_tokens)[:800] or 'none captured'}

Return ONLY JSON:
{{"per_category": {{"<category>": "what the winners claim here that {baseline} does NOT — be specific (name the exact keywords/claims), 1-2 sentences"}},
 "actions": ["5-7 precise, prioritised moves: e.g. 'Add a cooling/refreshing variant claim to sunscreen titles — 4 of top-10 lead with it, we don't'"],
 "summary": "3-4 sentences: the single biggest keyword/claim gap {baseline} must close to win the shelf"}}
JSON only."""
    raw = groq_chat(prompt, system="Output valid JSON only.", temperature=0.2, max_tokens=1600)
    try:
        return json.loads(raw[raw.index("{"): raw.rindex("}") + 1])
    except Exception:
        return {"per_category": {}, "actions": [], "summary": ""}


# ── orchestrator ──────────────────────────────────────────────────────────────
def build_shelf_report(listings, coverage, tracked_brands=None, progress_cb=None):
    platforms = sorted({l["platform"] for l in listings})
    keywords = sorted({l["keyword"] for l in listings})

    if progress_cb:
        progress_cb(1, 3, "Computing shelf share, search dominance, demand…")

    tracked = set(tracked_brands or []) | {BASELINE}

    sos_overall = share_of_shelf(listings)
    sos_by_platform = {p: share_of_shelf(listings, p) for p in platforms}
    dominance = search_dominance(listings)
    assortment = assortment_matrix(listings)
    demand = demand_and_price(listings)

    # distinctive keyword intelligence (category / competitor / product level).
    # exclude brand-name words so a brand's own name never counts as a "keyword".
    brand_stop = _brand_stop(listings)
    kw_by_cat = category_keywords(listings, categories=keywords, exclude=brand_stop)
    kw_by_comp = competitor_keywords(listings, tracked, exclude=brand_stop)
    kw_products = product_keywords(listings, exclude=brand_stop)

    if progress_cb:
        progress_cb(2, 3, "AI: keyword gaps vs competitors…")
    gaps_ai = gap_analysis(kw_by_cat, kw_by_comp, listings, exclude=brand_stop)

    # Deconstruct scorecard vs tracked competitors
    dom_by_brand = {d["brand"]: d for d in dominance}
    dem_by_brand = {d["brand"]: d for d in demand}
    sos_by_brand = {s["brand"]: s for s in sos_overall}
    scorecard = []
    for b in sorted(tracked):
        s = sos_by_brand.get(b, {})
        d = dom_by_brand.get(b, {})
        m = dem_by_brand.get(b, {})
        scorecard.append({
            "brand": b, "is_baseline": b == BASELINE,
            "shelf_pct": s.get("pct", 0.0), "skus": s.get("skus", 0),
            "top10": d.get("top10", 0), "top3": d.get("top3", 0),
            "avg_rank": d.get("avg_rank"),
            "total_reviews": m.get("total_reviews", 0), "avg_rating": m.get("avg_rating"),
            "avg_price": m.get("avg_price", 0), "price_tier": m.get("price_tier"),
            "platforms_present": sum(1 for p in platforms if assortment.get(b, {}).get(p)),
        })
    scorecard.sort(key=lambda x: -x["shelf_pct"])

    # distribution gaps for Deconstruct
    dec_present = assortment.get(BASELINE, {})
    gaps = [p for p in platforms if not dec_present.get(p)]

    methodology = _methodology(coverage, keywords, sorted(tracked))

    return {
        "platforms": platforms, "keywords": keywords, "coverage": coverage,
        "total_listings": len(listings), "tracked_brands": sorted(tracked),
        "share_of_shelf": sos_overall, "share_of_shelf_by_platform": sos_by_platform,
        "search_dominance": dominance, "assortment": assortment, "demand": demand,
        "keyword_by_category": kw_by_cat, "keyword_by_competitor": kw_by_comp,
        "product_keywords": kw_products, "gap_analysis": gaps_ai,
        "scorecard": scorecard, "deconstruct_distribution_gaps": gaps,
        "methodology": methodology,
    }


def _methodology(coverage, keywords, tracked):
    """Plain-English 'how each number is produced' — shown atop every sheet/section."""
    cov = " · ".join(f"{p}: {sum(c.values())} listings" for p, c in coverage.items())
    return {
        "overview": (
            "Every number below is read DIRECTLY from each marketplace's own listing data "
            "(no AI, no estimation). We scan the top listings on each platform's category "
            f"search — page 1, in the order the marketplace itself ranks them. "
            f"Categories scanned: {', '.join(keywords)}. Coverage this run — {cov}. "
            "Only the 'Keyword gap' commentary is AI-written, grounded in the scraped titles."),
        "share_of_shelf": (
            "Share of Shelf % = (a brand's SKUs appearing in the scanned top listings) ÷ "
            "(all scanned listings), per platform and blended. It measures VISIBILITY/presence "
            "on the shelf — not sales. Sorted high → low. Source: marketplace category-search results."),
        "search_dominance": (
            "Rank = the product's exact position in the platform's default category-search results "
            "(1 = very top). Top-3 / Top-10 = how many of a brand's listings land in positions 1–3 / "
            "1–10, summed across categories & platforms. Sorted by Top-10, then best average rank."),
        "assortment": (
            "SKU count per brand per platform (breadth of range on the shelf). A red/empty cell = "
            "that brand has no listing in the scanned top results on that platform = a distribution gap."),
        "demand": (
            "Total Reviews = sum of each brand's listing review counts (a PROXY for cumulative sales/"
            "traction — more reviews ≈ more sold over time; it is NOT a sales figure). Avg rating/price "
            "are means across the brand's scanned listings. Price tier: <₹350 budget · ₹350–700 mid · >₹700 premium."),
        "keywords": (
            "Keywords are DISTINCTIVE terms: frequent within a category's top listings but rare across "
            "other categories (TF-IDF), with the category's own query words removed — so you see 'gel, "
            "matte, brightening', not the obvious 'sunscreen'. Per-competitor/per-product views show which "
            "brand leads with which claims. The 'gap' read is AI, grounded in these scraped titles."),
        "all_listings": (
            "The raw evidence: every single listing scraped (platform · category · rank · brand · title · "
            "price · rating · reviews · URL). All other sheets are computed from exactly these rows — "
            "audit any number by filtering here. Flipkart price/rating are blank because Flipkart hides "
            "them from the fast fetch; those products ARE still counted for shelf-share, rank and keywords."),
        "tracked": "Tracked competitors (focus set): " + ", ".join(b for b in tracked if b != BASELINE),
    }


# ── Excel export (multi-sheet) ────────────────────────────────────────────────
def shelf_report_to_excel(report, listings):
    """Render the shelf report to a clean, team-readable multi-sheet workbook."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        BLACK = PatternFill("solid", fgColor="0A0A0A")
        LIME = PatternFill("solid", fgColor="C8F55A")
        GREY = PatternFill("solid", fgColor="F5F5F5")
        LIME_FONT = Font(bold=True, color="C8F55A", size=10)
        HDR = Alignment(horizontal="center", vertical="center", wrap_text=True)
        WRAP = Alignment(vertical="top", wrap_text=True)
        thin = Border(*(Side(style="thin", color="DDDDDD"),) * 4)

        def txt(v):
            if isinstance(v, list):
                return "\n".join(f"• {x}" for x in v)
            if isinstance(v, dict):
                return "\n".join(f"{k}: {val}" for k, val in v.items())
            return "" if v is None else v

        def head(ws, hs, row=1, h=34):
            for c, t in enumerate(hs, 1):
                cell = ws.cell(row=row, column=c, value=t)
                cell.fill = BLACK
                cell.font = LIME_FONT
                cell.alignment = HDR
                cell.border = thin
            ws.row_dimensions[row].height = h

        def widths(ws, ws_widths):
            for i, w in enumerate(ws_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def note(ws, text, ncols, row=1, height=46):
            """Methodology strip atop a sheet — 'how this is calculated', at a glance."""
            c = ws.cell(row=row, column=1, value="ⓘ  " + str(text))
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
            c.font = Font(italic=True, size=9, color="555555")
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.fill = PatternFill("solid", fgColor="F1F0EC")
            ws.row_dimensions[row].height = height

        meth = report.get("methodology", {})
        plats = report["platforms"]
        wb = openpyxl.Workbook()

        # 1 — Scorecard (headline). Row1 title · Row2 method note · Row3 header · Row4 data
        ws = wb.active
        assert ws is not None
        ws.title = "Scorecard"
        ws.cell(row=1, column=1, value="Decon AI — Digital Shelf & Market-Presence Scorecard "
                f"({report['total_listings']} listings · {len(report['keywords'])} categories · "
                f"{', '.join(plats)})").font = Font(bold=True, size=12)
        ws.merge_cells("A1:K1")
        note(ws, meth.get("overview", ""), 11, row=2, height=58)
        head(ws, ["Brand", "Shelf %", "SKUs", "Top-3", "Top-10", "Avg Rank",
                  "Total Reviews (proxy)", "Avg Rating", "Avg Price ₹", "Price Tier",
                  "Platforms /" + str(len(plats))], row=3)
        for i, s in enumerate(report["scorecard"], 4):
            vals = [s["brand"], s["shelf_pct"], s["skus"], s["top3"], s["top10"], s["avg_rank"],
                    s["total_reviews"], s["avg_rating"], s["avg_price"], s["price_tier"] or "—",
                    s["platforms_present"]]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.border = thin
                cell.alignment = WRAP if c == 1 else HDR
                if s["is_baseline"]:
                    cell.fill = LIME
                    cell.font = Font(bold=True)
                elif i % 2 == 0:
                    cell.fill = GREY
        ws.freeze_panes = "A4"
        widths(ws, [20, 9, 7, 7, 8, 9, 18, 10, 11, 10, 12])

        # 2 — Share of Shelf
        ws2 = wb.create_sheet("Share of Shelf")
        ncol = 3 + len(plats)
        note(ws2, meth.get("share_of_shelf", ""), ncol)
        head(ws2, ["Brand", "Overall %", "SKUs"] + [f"{p} %" for p in plats], row=2)
        sos_p = {p: {r["brand"]: r["pct"] for r in report["share_of_shelf_by_platform"].get(p, [])}
                 for p in plats}
        for i, s in enumerate(report["share_of_shelf"], 3):
            b = s["brand"]
            row = [b, s["pct"], s["skus"]] + [sos_p[p].get(b, 0.0) for p in plats]
            for c, v in enumerate(row, 1):
                cell = ws2.cell(row=i, column=c, value=v)
                cell.border = thin
                if b == BASELINE:
                    cell.fill = LIME
                    cell.font = Font(bold=True)
        ws2.freeze_panes = "A3"
        widths(ws2, [22, 10, 7] + [10] * len(plats))

        # 3 — Search Dominance
        ws3 = wb.create_sheet("Search Dominance")
        note(ws3, meth.get("search_dominance", ""), 5)
        head(ws3, ["Brand", "Top-3 appearances", "Top-10 appearances", "Avg Rank", "Listings seen"], row=2)
        for i, d in enumerate(report["search_dominance"][:40], 3):
            for c, v in enumerate([d["brand"], d["top3"], d["top10"], d["avg_rank"], d["listings"]], 1):
                cell = ws3.cell(row=i, column=c, value=v)
                cell.border = thin
                if d["brand"] == BASELINE:
                    cell.fill = LIME
                    cell.font = Font(bold=True)
        ws3.freeze_panes = "A3"
        widths(ws3, [22, 16, 16, 10, 12])

        # 4 — Assortment & Distribution
        ws4 = wb.create_sheet("Assortment")
        note(ws4, meth.get("assortment", ""), len(plats) + 2)
        head(ws4, ["Brand"] + plats + ["Total SKUs"], row=2)
        arows = sorted(report["assortment"].items(), key=lambda kv: -sum(kv[1].values()))
        for i, (b, pv) in enumerate(arows, 3):
            row = [b] + [pv.get(p, 0) for p in plats] + [sum(pv.values())]
            for c, v in enumerate(row, 1):
                cell = ws4.cell(row=i, column=c, value=v)
                cell.border = thin
                if b == BASELINE:
                    cell.fill = LIME
                    cell.font = Font(bold=True)
                elif c > 1 and v == 0:
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")  # gap = red
        ws4.freeze_panes = "A3"
        widths(ws4, [22] + [10] * len(plats) + [11])

        # 5 — Keywords by Category (distinctive terms)
        ws5 = wb.create_sheet("Keywords by Category")
        note(ws5, meth.get("keywords", ""), 2)
        head(ws5, ["Category", "Distinctive keywords the top listings use  —  keyword (× listings)"], row=2)
        for i, (cat, kws) in enumerate(report.get("keyword_by_category", {}).items(), 3):
            ws5.cell(row=i, column=1, value=cat).font = Font(bold=True)
            ws5.cell(row=i, column=2,
                     value=", ".join(f"{k['kw']} (×{k['listings']})" for k in kws)).alignment = WRAP
            for c in (1, 2):
                ws5.cell(row=i, column=c).border = thin
            ws5.row_dimensions[i].height = 46
        ws5.freeze_panes = "A3"
        widths(ws5, [24, 95])

        # 6 — Keywords by Competitor
        ws6 = wb.create_sheet("Keywords by Competitor")
        note(ws6, "Which distinctive claims/ingredients each tracked brand leads with across its "
             "listings — read as 'this brand owns these words'. keyword (× listings).", 2)
        head(ws6, ["Brand", "Keywords this brand leads with"], row=2)
        for i, (b, kws) in enumerate(report.get("keyword_by_competitor", {}).items(), 3):
            ws6.cell(row=i, column=1, value=b).font = Font(bold=True)
            ws6.cell(row=i, column=2,
                     value=", ".join(f"{k} (×{n})" for k, n in kws)).alignment = WRAP
            for c in (1, 2):
                ws6.cell(row=i, column=c).border = thin
            if b == BASELINE:
                ws6.cell(row=i, column=1).fill = LIME
            ws6.row_dimensions[i].height = 42
        ws6.freeze_panes = "A3"
        widths(ws6, [22, 95])

        # 7 — Product Keywords (product-wise bifurcation)
        ws7 = wb.create_sheet("Product Keywords")
        note(ws7, "Product-level breakdown of the standout listings — e.g. 'Dot & Key Watermelon "
             "Sunscreen → watermelon, cooling, hydrating'. So you can see exactly which words each "
             "top product leans on.", 6)
        head(ws7, ["Platform", "Category", "Rank", "Brand", "Product", "Keywords it uses"], row=2)
        for i, p in enumerate(report.get("product_keywords", []), 3):
            vals = [p["platform"], p["category"], p.get("rank"), p["brand"],
                    p["title"], ", ".join(p["keywords"])]
            for c, v in enumerate(vals, 1):
                cell = ws7.cell(row=i, column=c, value=v)
                cell.border = thin
                cell.alignment = WRAP
                if p["brand"] == BASELINE and c == 4:
                    cell.fill = LIME
            ws7.row_dimensions[i].height = 34
        ws7.freeze_panes = "A3"
        widths(ws7, [10, 16, 6, 16, 46, 40])

        # 8 — Keyword Gaps (AI): what winners claim that Deconstruct doesn't
        ws8 = wb.create_sheet("Keyword Gaps (AI)")
        ga = report.get("gap_analysis", {})
        ws8.cell(row=1, column=1, value="Keyword & claim gaps for Deconstruct (AI-analysed, grounded "
                 "in the scraped titles above)").font = Font(bold=True, size=12)
        ws8.merge_cells("A1:B1")
        ws8.cell(row=2, column=1, value=ga.get("summary", "")).alignment = WRAP
        ws8.merge_cells("A2:B2")
        ws8.row_dimensions[2].height = 60
        ws8.cell(row=4, column=1, value="Prioritised actions:").font = Font(bold=True)
        r = 5
        for a in ga.get("actions", []):
            ws8.cell(row=r, column=1, value=f"• {a}").alignment = WRAP
            ws8.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws8.row_dimensions[r].height = 30
            r += 1
        r += 1
        head(ws8, ["Category", "What the winners claim here that Deconstruct does NOT"], row=r)
        for cat, gap in (ga.get("per_category", {}) or {}).items():
            r += 1
            ws8.cell(row=r, column=1, value=cat).font = Font(bold=True)
            ws8.cell(row=r, column=2, value=txt(gap)).alignment = WRAP
            for c in (1, 2):
                ws8.cell(row=r, column=c).border = thin
            ws8.row_dimensions[r].height = 44
        widths(ws8, [22, 95])

        # 9 — All Listings (raw evidence)
        ws9 = wb.create_sheet("All Listings")
        note(ws9, meth.get("all_listings", ""), 11, height=58)
        head(ws9, ["Platform", "Category", "Rank", "Brand", "Tracked?", "Title",
                   "Price ₹", "MRP ₹", "Rating", "Reviews", "URL"], row=2)
        for i, l in enumerate(sorted(listings, key=lambda x: (x["platform"], x["keyword"], x["rank"])), 3):
            vals = [l["platform"], l["keyword"], l["rank"], l["brand"],
                    "Yes" if l.get("is_tracked") else "", l["title"],
                    int(l["price"]) or "", int(l["mrp"]) or "", l["rating"] or "",
                    l["reviews"] or "", l["url"]]
            for c, v in enumerate(vals, 1):
                cell = ws9.cell(row=i, column=c, value=v)
                cell.border = thin
                cell.alignment = WRAP
                if l["brand"] == BASELINE and c == 4:
                    cell.fill = LIME
        ws9.freeze_panes = "A3"
        widths(ws9, [10, 16, 6, 16, 9, 46, 9, 9, 8, 9, 40])

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()
    except Exception as e:
        print(f"[shelf] excel error: {e}")
        import traceback
        traceback.print_exc()
        return b""
