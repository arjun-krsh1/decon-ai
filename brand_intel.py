"""
brand_intel.py — the intelligence layer for the Brand Analyser.

Two halves:
  1. BRAND strategy — per brand: format mix (reel/carousel/static), cadence,
     engagement, themes, products, skin concerns, launches, new-market signals.
  2. INFLUENCER engine (the core of the tool) — for every influencer a brand
     collabs with, go one hop deeper: fetch their profile + recent posts and
     reason about WHO they are (sector/niche), their SIZE (tier), and WHAT they
     post (content style). Then roll up which tiers/sectors each competitor
     favours, who overlaps across brands, and where the whitespace is.

Ends with number-backed recommendations split into a Reels playbook, a
Static/Carousel playbook, and an Influencer strategy.

LLM calls go through llm.groq_chat (direct, cached-free, JSON) and fail soft.
"""

from __future__ import annotations

import json
from collections import Counter

from llm import groq_chat, groq_available
import brand_scraper as bs


# ── deterministic per-brand rollup ────────────────────────────────────────────
def _engagement(p):
    """Engagement score; reels get a modest boost from view count."""
    return (p.get("likes", 0) or 0) + (p.get("comments", 0) or 0) + (p.get("views", 0) or 0) // 100


def top_posts(posts, n=20):
    """The n highest-performing posts (what we analyse instead of everything)."""
    return sorted(posts, key=_engagement, reverse=True)[:n]


def _format_mix(posts):
    mix = Counter(p.get("type", "image") for p in posts)
    total = sum(mix.values()) or 1
    return ({k: mix.get(k, 0) for k in ("reel", "carousel", "image")},
            {k: round(100 * mix.get(k, 0) / total) for k in ("reel", "carousel", "image")})


def brand_rollup(brand, posts, is_baseline, days_back):
    reels = [p for p in posts if p.get("type") == "reel"]
    counts, pct = _format_mix(posts)
    avg_likes = round(sum(p.get("likes", 0) for p in posts) / len(posts)) if posts else 0
    avg_views = round(sum(p.get("views", 0) for p in reels) / len(reels)) if reels else 0
    top = max(posts, key=lambda p: p.get("likes", 0), default=None)
    return {
        "brand": brand,
        "is_baseline": is_baseline,
        "total_posts": len(posts),
        "format_counts": counts,
        "format_pct": pct,
        "avg_likes": avg_likes,
        "avg_reel_views": avg_views,
        "collabs": sum(1 for p in posts if p.get("is_collab")),
        "posts_per_week": round(len(posts) / max(days_back / 7, 1), 1),
        "top_post": {"type": top.get("type"), "likes": top.get("likes"),
                     "views": top.get("views"), "url": top.get("url"),
                     "caption": top.get("caption", "")[:160]} if top else {},
    }


# ── LLM: per-brand content strategy ───────────────────────────────────────────
def _parse_json(txt, default):
    """Extract the first JSON object/array from an LLM response, else `default`."""
    try:
        starts = [i for i in (txt.find("{"), txt.find("[")) if i >= 0]
        start = min(starts)
        end = max(txt.rfind("}"), txt.rfind("]")) + 1
        return json.loads(txt[start:end])
    except Exception:
        return default


def analyse_brand_strategy(brand, posts):
    if not groq_available() or not posts:
        return {"themes": [], "products_pushed": [], "skin_concerns": [],
                "launches": [], "new_market_signals": [], "top_hooks": [],
                "content_summary": "AI unavailable — connect Groq for strategy read."}
    lines = "\n".join(
        f"- [{p.get('type')}, {p.get('likes',0)}❤ {p.get('views',0)}▶{' COLLAB' if p.get('is_collab') else ''}] "
        f"{p.get('caption','')[:200]} #{' #'.join(p.get('hashtags',[])[:6])}"
        for p in posts[:25])
    prompt = f"""You are a social-media competitive analyst. Analyse {brand}'s last-20-days Instagram posts.

POSTS:
{lines}

Return ONLY JSON:
{{"themes": ["recurring content themes"],
 "products_pushed": ["specific products/ranges spotlighted"],
 "skin_concerns": ["concerns targeted e.g. acne, pigmentation"],
 "launches": ["new product/campaign launches you can detect, else empty"],
 "new_market_signals": ["signals of new audiences/markets: regional, male, gen-z, festival, new category — else empty"],
 "top_hooks": ["the strongest hooks/angles they use"],
 "content_summary": "3-4 sentences: what {brand} is doing on IG right now and where they seem headed"}}
JSON only."""
    return _parse_json(groq_chat(prompt, system="Output valid JSON only.",
                                 temperature=0.2, max_tokens=900),
                       {"themes": [], "products_pushed": [], "skin_concerns": [],
                        "launches": [], "new_market_signals": [], "top_hooks": [],
                        "content_summary": ""})


# ── LLM: influencer characterisation (the core engine) ────────────────────────
def characterise_influencers(profiles, posts_by_influencer):
    """profiles: {handle: {followers,tier,category,full_name}}. Returns {handle: {...}}."""
    out = {}
    if not profiles:
        return out
    items = list(profiles.items())
    if not groq_available():
        for h, p in items:
            out[h] = {"sector": p.get("category", "") or "unknown", "content_style": "",
                      "audience_niche": ""}
        return out

    # batch ~8 influencers per call to keep prompts tight
    for i in range(0, len(items), 8):
        batch = items[i:i + 8]
        blocks = []
        for h, p in batch:
            caps = " | ".join(x.get("caption", "")[:90] for x in posts_by_influencer.get(h, [])[:5])
            blocks.append(f'@{h} — {p.get("followers",0)} followers ({p.get("tier","")}), '
                          f'IG category: {p.get("category","")}. Recent posts: {caps or "n/a"}')
        prompt = f"""Classify each Instagram creator for a skincare brand's competitor analysis.

CREATORS:
{chr(10).join(blocks)}

For EACH, return an object in a JSON array (same order):
{{"handle": "without @",
  "sector": "their primary niche: dermatologist/skincare-creator/beauty/makeup/lifestyle/fashion/fitness/mom/comedy/food/other",
  "content_style": "how they make content: educational/GRWM/reviews/aspirational/comedy/tutorials/UGC",
  "audience_niche": "who follows them, one short phrase"}}
JSON array only."""
        arr = _parse_json(groq_chat(prompt, system="Output a valid JSON array only.",
                                    temperature=0.2, max_tokens=1200), [])
        for obj in arr if isinstance(arr, list) else []:
            h = str(obj.get("handle", "")).lstrip("@")
            if h:
                out[h] = {"sector": obj.get("sector", ""), "content_style": obj.get("content_style", ""),
                          "audience_niche": obj.get("audience_niche", "")}
    # ensure every profile has an entry
    for h, p in items:
        out.setdefault(h, {"sector": p.get("category", "") or "unknown",
                           "content_style": "", "audience_niche": ""})
    return out


# ── cross-competitor trends + whitespace ──────────────────────────────────────
def _cross_trends(rollups, strategies):
    concern_freq, ing_freq, theme_freq = Counter(), Counter(), Counter()
    fmt = Counter()
    for r in rollups:
        for k, v in r["format_counts"].items():
            fmt[k] += v
    for s in strategies.values():
        for c in s.get("skin_concerns", []):
            concern_freq[str(c).lower().strip()] += 1
        for t in s.get("themes", []):
            theme_freq[str(t).lower().strip()] += 1
    return {"format_totals": dict(fmt),
            "top_concerns": concern_freq.most_common(8),
            "top_themes": theme_freq.most_common(8)}


# ── recommendations (Reels / Static / Influencer) ─────────────────────────────
def _recommend(report):
    if not groq_available():
        return {"reels_playbook": [], "static_playbook": [], "influencer_strategy": [],
                "summary": "Connect Groq for AI recommendations."}
    dec = next((b for b in report["brands"] if b.get("is_baseline")), {})
    comp = [b for b in report["brands"] if not b.get("is_baseline")]
    comp_lines = "\n".join(
        f"- {b['brand']}: {b['format_pct'].get('reel',0)}% reels / {b['format_pct'].get('carousel',0)}% carousel, "
        f"{b['posts_per_week']}/wk, avg {b['avg_likes']}❤, {b['collabs']} collabs, "
        f"tiers={b.get('influencer_tier_mix',{})}" for b in comp)
    infl = report.get("influencer_analysis", {})
    prompt = f"""You are Deconstruct's head of social strategy. Based on 20-day Instagram competitor data, give an actionable plan.

DECONSTRUCT (baseline): {dec.get('format_pct',{})} formats, {dec.get('posts_per_week',0)}/wk, avg {dec.get('avg_likes',0)} likes, {dec.get('collabs',0)} collabs.
COMPETITORS:
{comp_lines}
INFLUENCER TIER USAGE (across competitors): {infl.get('tier_distribution',{})}
TOP INFLUENCER SECTORS: {infl.get('sector_distribution',{})}
CROSS TRENDS: concerns={report.get('trends',{}).get('top_concerns',[])[:5]}

Return ONLY JSON:
{{"reels_playbook": ["4-5 specific Reel plays Deconstruct should run, grounded in what's working for competitors"],
 "static_playbook": ["4-5 specific static/carousel plays (education, claims, comparisons)"],
 "influencer_strategy": ["4-5 moves: which tiers (nano/micro/macro) and sectors to target, and why, vs competitors"],
 "summary": "3-4 sentence executive read of the biggest opportunity for Deconstruct"}}
JSON only."""
    return _parse_json(groq_chat(prompt, system="Output valid JSON only.",
                                 temperature=0.3, max_tokens=1600),
                       {"reels_playbook": [], "static_playbook": [], "influencer_strategy": [],
                        "summary": ""})


# ── orchestrator ──────────────────────────────────────────────────────────────
def build_report(by_brand, days_back=25, infl_cap_per_brand=8,
                 include_hashtag_creators=True, top_n_posts=20,
                 min_followers=10000, progress_cb=None):
    """Full brand-media report. `by_brand` = {brand: [posts]} from brand_scraper."""
    baseline = {str(b["name"]): bool(b["is_baseline"]) for b in bs.BRANDS}
    brands_order = [str(b["name"]) for b in bs.BRANDS if str(b["name"]) in by_brand]

    # 1. per-brand rollups + strategy
    rollups, strategies = [], {}
    for i, brand in enumerate(brands_order):
        if progress_cb:
            progress_cb(i, len(brands_order) + 2, f"Analysing {brand}'s top posts…")
        all_posts = by_brand.get(brand, [])
        top = top_posts(all_posts, top_n_posts)   # focus on high performers only
        r = brand_rollup(brand, top, baseline.get(brand, False), days_back)
        r["total_posts"] = len(all_posts)          # true volume (context)
        r["analysed_posts"] = len(top)             # how many we actually studied
        r["posts_per_week"] = round(len(all_posts) / max(days_back / 7, 1), 1)  # true cadence
        rollups.append(r)
        strategies[brand] = analyse_brand_strategy(brand, top)

    # 2. influencer engine — discover creators from BOTH brackets, go one hop deeper
    #    (a) featured on the brand's own feed  (b) creators who hashtag the brand
    #        (the majority bracket) — then profile + characterise each.
    if progress_cb:
        progress_cb(len(brands_order), len(brands_order) + 2, "Discovering & profiling creators…")
    featured = bs.collab_influencers(by_brand)
    hashtagged = bs.fetch_hashtag_creators(days_back) if include_hashtag_creators else {}
    capped = {}
    for b in brands_order:
        htag = [h for h, _ in hashtagged.get(b, [])]      # majority bracket, ranked
        merged = list(dict.fromkeys(htag + featured.get(b, [])))
        capped[b] = merged[:infl_cap_per_brand]
    all_influencers = list(dict.fromkeys(h for hs in capped.values() for h in hs))
    profiles = bs.fetch_influencer_profiles(all_influencers)
    # DROP tiny/vague accounts a brand would never collab with (follower floor).
    profiles = {h: p for h, p in profiles.items()
                if (p.get("followers", 0) or 0) >= min_followers}
    # fetch posts + characterise ONLY the collab-worthy survivors (saves credits)
    infl_posts = bs.fetch_influencer_posts(list(profiles.keys()))
    characterised = characterise_influencers(profiles, infl_posts)

    # merge profile + characterisation
    influencers = {}
    for h, prof in profiles.items():
        influencers[h] = {**prof, **characterised.get(h, {})}

    # attach per-brand influencer roster + tier mix
    for r in rollups:
        roster = []
        tier_mix = Counter()
        for h in capped.get(r["brand"], []):
            inf = influencers.get(h)
            if inf:
                roster.append({"handle": h, **inf})
                tier_mix[inf.get("tier", "?")] += 1
        r["influencers"] = roster
        r["influencer_tier_mix"] = dict(tier_mix)
        r["strategy"] = strategies.get(r["brand"], {})

    # 3. cross trends + influencer distributions + overlaps
    tier_dist = Counter(i.get("tier", "?") for i in influencers.values())
    sector_dist = Counter(i.get("sector", "?") for i in influencers.values() if i.get("sector"))
    brand_by_infl = {}
    for b, hs in capped.items():
        for h in hs:
            if h in influencers:   # only surviving (collab-worthy) creators
                brand_by_infl.setdefault(h, []).append(b)
    overlaps = {h: bl for h, bl in brand_by_infl.items() if len(bl) > 1}
    for h, inf in influencers.items():   # tag each creator with the brand(s) they post about
        inf["brands"] = brand_by_infl.get(h, [])

    report = {
        "generated_at": None,  # stamped by caller
        "days_back": days_back,
        "brands": rollups,
        "trends": _cross_trends(rollups, strategies),
        "influencers": influencers,
        "influencer_analysis": {
            "tier_distribution": dict(tier_dist),
            "sector_distribution": dict(sector_dist.most_common(10)),
            "overlaps": overlaps,   # influencers working with 2+ competitors
            "total_influencers": len(influencers),
        },
    }
    if progress_cb:
        progress_cb(len(brands_order) + 1, len(brands_order) + 2, "Writing recommendations…")
    report["recommendations"] = _recommend(report)
    return report


# ── Excel export (multi-sheet) ────────────────────────────────────────────────
def report_to_excel(report):
    """Render a brand-media report to a multi-sheet Excel workbook (bytes)."""
    try:
        import io
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        BLACK = PatternFill("solid", fgColor="0A0A0A")
        LIME = PatternFill("solid", fgColor="C8F55A")
        LIME_FONT = Font(bold=True, color="C8F55A", size=10)
        HDR = Alignment(horizontal="center", vertical="center", wrap_text=True)
        WRAP = Alignment(vertical="top", wrap_text=True)
        thin = Border(*(Side(style="thin", color="DDDDDD"),) * 4)

        def txt(v):
            if isinstance(v, list):
                return "\n".join(f"• {x}" for x in v)
            if isinstance(v, dict):
                return "; ".join(f"{k}: {val}" for k, val in v.items())
            return "" if v is None else v

        def headers(ws, hs, row=1, h=40):
            for c, t in enumerate(hs, 1):
                cell = ws.cell(row=row, column=c, value=t)
                cell.fill = BLACK
                cell.font = LIME_FONT
                cell.alignment = HDR
                cell.border = thin
            ws.row_dimensions[row].height = h

        wb = openpyxl.Workbook()
        brands = report.get("brands", [])
        influencers = report.get("influencers", {})
        recs = report.get("recommendations", {})
        ia = report.get("influencer_analysis", {})

        # Sheet 1 — Executive Summary + playbooks
        ws = wb.active
        assert ws is not None
        ws.title = "Summary & Playbooks"
        ws.cell(row=1, column=1, value="Decon AI — Brand Media Analysis (Instagram, last "
                f"{report.get('days_back',20)} days)").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=recs.get("summary", "")).alignment = WRAP
        ws.merge_cells("A2:E2")
        ws.row_dimensions[2].height = 60
        row = 4
        for title, key in [("🎬 REELS PLAYBOOK", "reels_playbook"),
                           ("🖼️ STATIC / CAROUSEL PLAYBOOK", "static_playbook"),
                           ("🤝 INFLUENCER STRATEGY", "influencer_strategy")]:
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color="0A0A0A")
            row += 1
            for item in recs.get(key, []):
                c = ws.cell(row=row, column=1, value=f"• {item}")
                c.alignment = WRAP
                c.fill = LIME if False else PatternFill("solid", fgColor="F4FAE9")
                ws.row_dimensions[row].height = 30
                row += 1
            row += 1
        ws.column_dimensions["A"].width = 100

        # Sheet 2 — Brand Strategy (formats, engagement, launches)
        ws2 = wb.create_sheet("Brand Strategy")
        headers(ws2, ["Brand", "Role", "Posts", "Reel %", "Carousel %", "Static %",
                      "Posts/Wk", "Avg Likes", "Avg Reel Views", "Collabs",
                      "Themes", "Products Pushed", "Skin Concerns",
                      "Launches Detected", "New-Market Signals", "Strategy Summary"])
        for i, b in enumerate(brands, 2):
            s = b.get("strategy", {})
            base = b.get("is_baseline")
            vals = [b["brand"], "⭐ Baseline" if base else "Competitor", b["total_posts"],
                    b["format_pct"].get("reel", 0), b["format_pct"].get("carousel", 0),
                    b["format_pct"].get("image", 0), b["posts_per_week"], b["avg_likes"],
                    b["avg_reel_views"], b["collabs"],
                    txt(s.get("themes", [])), txt(s.get("products_pushed", [])),
                    txt(s.get("skin_concerns", [])), txt(s.get("launches", [])),
                    txt(s.get("new_market_signals", [])), s.get("content_summary", "")]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=i, column=c, value=v)
                cell.border = thin
                cell.alignment = WRAP
                if c == 1:
                    cell.fill = LIME if base else PatternFill("solid", fgColor="F5F5F5")
                    cell.font = Font(bold=True)
            ws2.row_dimensions[i].height = 110
        for i, w in enumerate([16, 12, 7, 7, 9, 8, 8, 9, 12, 8, 30, 30, 24, 28, 30, 55], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A2"

        # Sheet 3 — Influencers (the core deliverable)
        ws3 = wb.create_sheet("Influencers")
        headers(ws3, ["Handle", "Works With (brand)", "Tier", "Followers",
                      "Sector / Niche", "Content Style", "Audience"])
        rows = sorted(influencers.items(),
                      key=lambda kv: kv[1].get("followers", 0) or 0, reverse=True)
        for i, (h, inf) in enumerate(rows, 2):
            vals = ["@" + h, txt(inf.get("brands", [])), inf.get("tier", ""),
                    inf.get("followers", 0), inf.get("sector", ""),
                    inf.get("content_style", ""), inf.get("audience_niche", "")]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=i, column=c, value=v)
                cell.border = thin
                cell.alignment = WRAP
                if c == 2 and len(inf.get("brands", [])) > 1:  # cross-brand creator
                    cell.fill = PatternFill("solid", fgColor="FBF3EF")
            ws3.row_dimensions[i].height = 30
        for i, w in enumerate([22, 24, 9, 12, 22, 22, 30], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w
        ws3.freeze_panes = "A2"

        # Sheet 4 — Trends & Whitespace
        ws4 = wb.create_sheet("Trends & Whitespace")
        ws4.cell(row=1, column=1, value="Cross-Competitor Trends").font = Font(bold=True, size=13)
        tr = report.get("trends", {})
        r = 3
        for label, data in [("Format totals", tr.get("format_totals", {})),
                            ("Top skin concerns", dict(tr.get("top_concerns", []))),
                            ("Top themes", dict(tr.get("top_themes", []))),
                            ("Influencer tier distribution", ia.get("tier_distribution", {})),
                            ("Influencer sector distribution", dict(ia.get("sector_distribution", {}))),
                            ("Cross-brand creators (overlaps)", {k: ", ".join(v) for k, v in ia.get("overlaps", {}).items()})]:
            ws4.cell(row=r, column=1, value=label).font = Font(bold=True, size=11)
            r += 1
            for k, v in (data.items() if isinstance(data, dict) else []):
                ws4.cell(row=r, column=1, value=str(k)).border = thin
                ws4.cell(row=r, column=2, value=txt(v)).border = thin
                r += 1
            r += 1
        ws4.column_dimensions["A"].width = 34
        ws4.column_dimensions["B"].width = 50

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()
    except Exception as e:
        print(f"[brand] excel error: {e}")
        import traceback
        traceback.print_exc()
        return b""
